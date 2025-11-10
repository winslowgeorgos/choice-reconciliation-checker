
import streamlit as st
from io import BytesIO
import pandas as pd
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go

# Import functions from other pages (assuming these files exist in the same directory)
from fx_reconcilliation_app_page import fx_reconciliation_app
from fx_trade_reconciliation_page import graphed_analysis_app
from intermediary_bank_reconciliation_page import intermediary_bank_reconciliation_app
from interfund_bank_reconciliation_page import interfund_bank_reconciliation_app
from intermediary_bank_reconciliation_page import intermediary_bank_reconciliation_app
from combine_match_results_page import run_cross_match_analysis, cross_match_analysis_app
from business_fx_reconciliation_page import business_reconciliation_app   # NEW

import requests
import time


import re
from typing import Optional, Dict
import json


# --- FastForex API Configuration ---
FASTFOREX_API_URL = "https://api.fastforex.io/fetch-all"
FASTFOREX_API_KEY = "4b744777d6-9c3eed3143-t4gxsb"
CACHE_FILE = "exchange_rates_cache.json"

# --- Target Currencies ---
TARGET_CURRENCIES = ["KES", "USD", "EUR", "GBP", "CNY", "UGX", "RWF", "TZS", "ZAR"]

# --- Currency Code Mapping ---
CURRENCY_NAME_MAP = {
    "KES": "KENYA SHILLING",
    "USD": "US DOLLAR",
    "GBP": "STG POUND", 
    "EUR": "EURO",
    "CNY": "CHINESE YUAN",
    "UGX": "UGANDA SHILLING",
    "RWF": "RWANDA FRANC",
    "TZS": "TANZANIA SHILLING",
    "ZAR": "SA RAND"
}

def get_live_exchange_rates(base_currency: str = "KES") -> Optional[Dict[str, float]]:
    """
    Fetch live exchange rates from FastForex API and return rates for converting TO KES.
    Returns dict with currency codes as keys and conversion rates TO KES as values.
    """
    try:
        print("🌍 Fetching live exchange rates from FastForex API...")
        
        params = {
            "from": base_currency,
            "api_key": FASTFOREX_API_KEY
        }

        headers = {
            "accept": "application/json",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }

        response = requests.get(FASTFOREX_API_URL, params=params, headers=headers, timeout=30)
        
        if response.status_code != 200:
            st.error(f"❌ API returned status code: {response.status_code}")
            return None

        data = response.json()
        
        if "results" not in data:
            st.error("❌ Invalid API response format")
            return None

        results = data.get('results', {})
        
        # Convert rates to KES (inverse of the provided rates)
        # FastForex gives: 1 KES = X FOREIGN, but we need: 1 FOREIGN = Y KES
        exchange_rates_to_kes = {}
        
        for currency_code, rate_from_kes in results.items():
            if currency_code in TARGET_CURRENCIES and rate_from_kes > 0:
                # Convert to KES: 1 FOREIGN = 1 / rate KES
                rate_to_kes = 1.0 / rate_from_kes
                exchange_rates_to_kes[currency_code] = rate_to_kes
                print(f"✅ {currency_code}: 1 {currency_code} = {rate_to_kes:.2f} KES")
        
        # Always include KES with rate 1
        exchange_rates_to_kes["KES"] = 1.0
        
        st.success(f"✅ Fetched {len(exchange_rates_to_kes)} exchange rates")
        return exchange_rates_to_kes
        
    except requests.exceptions.RequestException as e:
        st.error(f"❌ Network error fetching exchange rates: {e}")
        return get_fallback_rates()
    except json.JSONDecodeError:
        st.error("❌ Invalid JSON response from exchange rate API")
        return get_fallback_rates()
    except Exception as e:
        st.error(f"❌ Unexpected error fetching exchange rates: {e}")
        return get_fallback_rates()

def get_fallback_rates() -> Dict[str, float]:
    """Provide fallback exchange rates when API fails."""
    st.warning("⚠️ Using fallback exchange rates")
    
    # Fallback rates (1 FOREIGN = X KES)
    fallback_rates = {
        "KES": 1.0,
        "USD": 129.24,    # 1 USD = 129.24 KES
        "EUR": 150.64,    # 1 EUR = 150.64 KES
        "GBP": 173.38,    # 1 GBP = 173.38 KES
        "CNY": 18.15,     # 1 CNY = 18.15 KES
        "UGX": 0.037,     # 1 UGX = 0.037 KES
        "RWF": 0.089,     # 1 RWF = 0.089 KES
        "TZS": 0.053,     # 1 TZS = 0.053 KES
        "ZAR": 7.45       # 1 ZAR = 7.45 KES
    }
    
    return fallback_rates

def convert_to_kes(amount: float, currency: str, exchange_rates: Dict[str, float]) -> float:
    """
    Convert amount from given currency to KES.
    
    Args:
        amount: Amount in original currency
        currency: Currency code (USD, EUR, etc.)
        exchange_rates: Dict with conversion rates TO KES
    
    Returns:
        Amount converted to KES
    """
    if not amount or pd.isna(amount):
        return 0.0
    
    currency = currency.upper().strip()
    
    # Handle special currency variants
    currency_merge_map = {
        "KES-SPECIAL": "KES",
        "USD-SPECIAL": "USD",
        "USD-DCD": "USD",
        "EUR-SPECIAL": "EUR",
        "GBP-SPECIAL": "GBP",
    }
    currency = currency_merge_map.get(currency, currency)
    
    if currency == "KES":
        return float(amount)
    
    if currency in exchange_rates:
        rate = exchange_rates[currency]
        return float(amount) * rate
    else:
        st.warning(f"⚠️ No exchange rate found for {currency}, using 1:1 conversion")
        return float(amount)

# --- Main reporting function (your existing code updated) ---

def generate_cash_summary_report(per_bank_df: pd.DataFrame):
    """Generate the cash summary report with currency conversion to KES."""
    
    if not per_bank_df.empty:
        # --- Normalize column names ---
        per_bank_df.columns = (
            per_bank_df.columns.str.strip()
            .str.replace(" ", "_")
            .str.lower()
        )

        # --- Ensure required columns exist ---
        required_cols = ["currency", "bank", "opening_balance", "closing_balance"]
        missing = [c for c in required_cols if c not in per_bank_df.columns]
        if missing:
            st.warning(f"⚠️ Missing required columns: {missing}")
        else:
            # --- Combine SPECIAL and DCD currency variants ---
            currency_merge_map = {
                "KES-SPECIAL": "KES",
                "USD-SPECIAL": "USD",
                "USD-DCD": "USD",
                "EUR-SPECIAL": "EUR",
                "GBP-SPECIAL": "GBP",
            }
            per_bank_df["currency"] = (
                per_bank_df["currency"].str.upper().replace(currency_merge_map)
            )

            # --- Get live exchange rates ---
            st.info("🔄 Fetching live exchange rates...")
            exchange_rates = get_live_exchange_rates("KES")
            
            # Display current rates for transparency
            if exchange_rates:
                # Show rates in both directions for clarity
                rate_info_kes_to_foreign = " | ".join([f"1 KES = {1/rate:.4f} {curr}" 
                                    for curr, rate in exchange_rates.items() 
                                    if curr in ['USD', 'EUR', 'GBP'] and curr != "KES"])
                
                rate_info_foreign_to_kes = " | ".join([f"1 {curr} = {rate:.2f} KES" 
                                    for curr, rate in exchange_rates.items() 
                                    if curr in ['USD', 'EUR', 'GBP'] and curr != "KES"])
                
                st.caption(f"💱 Live Rates (KES to Foreign): {rate_info_kes_to_foreign}")
                st.caption(f"💱 Live Rates (Foreign to KES): {rate_info_foreign_to_kes}")

            # --- Compute currency summary automatically ---
            currency_summary = (
                per_bank_df.groupby("currency", as_index=False)[["opening_balance", "closing_balance"]]
                .sum()
                .sort_values("currency")
            )

            # --- Add KES conversion columns ---
            if exchange_rates:
                currency_summary['opening_balance_kes'] = currency_summary.apply(
                    lambda x: convert_to_kes(x['opening_balance'], x['currency'], exchange_rates), 
                    axis=1
                )
                currency_summary['closing_balance_kes'] = currency_summary.apply(
                    lambda x: convert_to_kes(x['closing_balance'], x['currency'], exchange_rates), 
                    axis=1
                )

            # --- Add Grand Total row ---
            grand_total_data = {
                "currency": "GRAND TOTAL",
                "opening_balance": currency_summary["opening_balance"].sum(),
                "closing_balance": currency_summary["closing_balance"].sum()
            }
            
            # Add KES totals for grand total if conversion columns exist
            if 'opening_balance_kes' in currency_summary.columns:
                grand_total_data["opening_balance_kes"] = currency_summary["opening_balance_kes"].sum()
                grand_total_data["closing_balance_kes"] = currency_summary["closing_balance_kes"].sum()

            grand_total = pd.DataFrame([grand_total_data])
            currency_summary = pd.concat([currency_summary, grand_total], ignore_index=True)

            # === Create Bank Consolidated Summary (KES) ===
            # Derive clean bank names (strip currency suffixes)
            per_bank_df["bank_clean"] = (
                per_bank_df["bank"]
                .astype(str)
                .str.replace(r"\b(USD|EUR|GBP|KES|CNY|ZAR|TZS|UGX|RWF)\b", "", regex=True)
                .str.replace(r"[-_/]+$", "", regex=True)
                .str.strip()
            )

            # Convert all balances to KES equivalent
            if exchange_rates:
                per_bank_df["fx_rate_to_KES"] = per_bank_df["currency"].map(exchange_rates).fillna(1.0)
                per_bank_df["opening_balance_KES"] = per_bank_df["opening_balance"] * per_bank_df["fx_rate_to_KES"]
                per_bank_df["closing_balance_KES"] = per_bank_df["closing_balance"] * per_bank_df["fx_rate_to_KES"]
            else:
                # Fallback to default rates if live rates not available
                fx_rates = {
                    "KES": 1.0,
                    "USD": 130.0,
                    "EUR": 140.0,
                    "GBP": 160.0,
                }
                per_bank_df["fx_rate_to_KES"] = per_bank_df["currency"].map(fx_rates).fillna(1.0)
                per_bank_df["opening_balance_KES"] = per_bank_df["opening_balance"] * per_bank_df["fx_rate_to_KES"]
                per_bank_df["closing_balance_KES"] = per_bank_df["closing_balance"] * per_bank_df["fx_rate_to_KES"]

            # Consolidate by clean bank name
            bank_summary_kes = (
                per_bank_df.groupby("bank_clean", as_index=False)[["opening_balance_KES", "closing_balance_KES"]]
                .sum()
                .sort_values("closing_balance_KES", ascending=False)
            )

            # Identify top 3 banks by closing balance for highlighting
            top_banks = bank_summary_kes.nlargest(3, "closing_balance_KES")
            top_bank_names = top_banks["bank_clean"].tolist()

            # === Create Excel workbook with XlsxWriter ===
            excel_buffer = BytesIO()
            wb = Workbook(write_only=False)
            ws = wb.active
            ws.title = "Cash Summary"

            # --- Create custom number formats ---
            kes_currency_format = '#,##0.00" KSh"'
            number_format = '#,##0.00'
            
            # Store formats for reuse
            formats = {
                'kes_currency': kes_currency_format,
                'number': number_format,
                'header_bold': Font(bold=True, size=14),
                'section_header': Font(bold=True),
                'column_header': Font(bold=True),
                'grand_total': Font(bold=True, color="FFFFFF"),
            }

            # --- Header ---
            report_date = pd.Timestamp.today().strftime("%d %B %Y").upper()
            ws.merge_cells("A1:L1")
            ws["A1"] = f"CASH SUMMARY AS AT {report_date}"
            ws["A1"].font = formats['header_bold']
            ws["A1"].alignment = Alignment(horizontal="center")

            # --- Define section order ---
            currency_order = ["KES", "USD", "EUR", "GBP", "CNY", "UGX", "RWF", "TZS", "ZAR"]
            start_col = 1  # Excel columns start at A (1)

            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            # --- Write section headers for each currency ---
            for i, currency in enumerate(currency_order):
                df_cur = per_bank_df[per_bank_df["currency"].str.upper() == currency]
                if df_cur.empty:
                    continue

                # Column offsets (3 columns per currency block)
                col_offset = (i * 3) + start_col

                # Section title
                ws.merge_cells(
                    start_row=3, start_column=col_offset,
                    end_row=3, end_column=col_offset + 2
                )
                ws.cell(row=3, column=col_offset).value = f"BANK {currency} ACCOUNTS"
                ws.cell(row=3, column=col_offset).font = formats['section_header']
                ws.cell(row=3, column=col_offset).alignment = Alignment(horizontal="center")

                # Column headers
                headers = ["BANK NAME", "OPENING BALANCE", "CLOSING BALANCE"]
                for j, header in enumerate(headers):
                    cell = ws.cell(row=4, column=col_offset + j, value=header)
                    cell.font = formats['column_header']
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

                # Write data rows
                for r_idx, row in enumerate(df_cur.itertuples(index=False), start=5):
                    row_dict = row._asdict()
                    bank_name = row_dict.get("bank", "")
                    opening = row_dict.get("opening_balance", 0)
                    closing = row_dict.get("closing_balance", 0)

                    ws.cell(row=r_idx, column=col_offset, value=bank_name)
                    ws.cell(row=r_idx, column=col_offset + 1, value=opening)
                    ws.cell(row=r_idx, column=col_offset + 2, value=closing)

                    # Apply formatting and borders
                    for j in range(3):
                        cell = ws.cell(row=r_idx, column=col_offset + j)
                        cell.border = thin_border
                        if j > 0:
                            cell.number_format = formats['number']
                            cell.alignment = Alignment(horizontal="right")

            # === Add Bank Consolidated Summary (KES) on the same sheet ===
            consolidated_start_row = ws.max_row + 3
            
            # Consolidated Summary Header
            ws.merge_cells(
                start_row=consolidated_start_row,
                start_column=1,
                end_row=consolidated_start_row,
                end_column=5
            )
            ws.cell(row=consolidated_start_row, column=1, value="BANK CONSOLIDATED SUMMARY (KES)").font = Font(bold=True, size=12)
            
            # Consolidated Summary Column Headers
            consolidated_headers = ["BANK", "OPENING BALANCE (KES)", "CLOSING BALANCE (KES)", "CHANGE (KES)", "GROWTH %"]
            for j, header in enumerate(consolidated_headers):
                cell = ws.cell(row=consolidated_start_row + 1, column=j + 1, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            # Consolidated Summary Data
            for r_idx, row in enumerate(bank_summary_kes.itertuples(index=False), start=consolidated_start_row + 2):
                bank_name = row.bank_clean
                opening_kes = row.opening_balance_KES
                closing_kes = row.closing_balance_KES
                change_kes = closing_kes - opening_kes
                growth_pct = (change_kes / opening_kes * 100) if opening_kes != 0 else 0

                # Write data
                ws.cell(row=r_idx, column=1, value=bank_name)
                ws.cell(row=r_idx, column=2, value=opening_kes)
                ws.cell(row=r_idx, column=3, value=closing_kes)
                ws.cell(row=r_idx, column=4, value=change_kes)
                ws.cell(row=r_idx, column=5, value=growth_pct)

                # Apply formatting
                for c in range(1, 6):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.border = thin_border
                    
                    if c in [2, 3, 4]:  # Currency columns
                        cell.number_format = formats['kes_currency']
                        cell.alignment = Alignment(horizontal="right")
                    elif c == 5:  # Percentage column
                        cell.number_format = "0.00%"
                        cell.alignment = Alignment(horizontal="right")
                    
                    # Highlight top 3 banks
                    if bank_name in top_bank_names:
                        if bank_name == top_bank_names[0]:  # Top bank - green
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        else:  # Other top 3 banks - blue
                            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

            # === Totals Section (after consolidated summary) ===
            total_row_start = ws.max_row + 2
            
            # Add exchange rate info
            if exchange_rates:
                # Show rates for conversion to KES (more relevant for the report)
                rate_display = []
                for curr in ['USD', 'EUR', 'GBP', 'CNY', 'ZAR']:
                    if curr in exchange_rates and curr != "KES":
                        rate_display.append(f"1 {curr} = {exchange_rates[curr]:.2f} KES")
                
                if rate_display:
                    ws.merge_cells(
                        start_row=total_row_start,
                        start_column=1,
                        end_row=total_row_start,
                        end_column=6
                    )
                    rate_text = "Exchange Rates (to KES): " + " | ".join(rate_display)
                    ws.cell(row=total_row_start, column=1, value=rate_text).font = Font(italic=True, size=9)
                    total_row_start += 1

            ws.merge_cells(
                start_row=total_row_start,
                start_column=1,
                end_row=total_row_start,
                end_column=6
            )
            ws.cell(row=total_row_start, column=1, value="TOTALS BY CURRENCY").font = Font(bold=True, size=12)

            # Totals headers - expanded to include KES columns
            totals_headers = ["CURRENCY", "OPENING TOTAL", "CLOSING TOTAL"]
            if 'opening_balance_kes' in currency_summary.columns:
                totals_headers.extend(["OPENING (KES)", "CLOSING (KES)"])

            for j, header in enumerate(totals_headers):
                cell = ws.cell(row=total_row_start + 1, column=j + 1, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            # Totals data
            for r_idx, row in enumerate(currency_summary.itertuples(index=False), start=total_row_start + 2):
                row_dict = row._asdict()
                currency = row_dict.get("currency", "")
                opening_total = row_dict.get("opening_balance", 0)
                closing_total = row_dict.get("closing_balance", 0)

                ws.cell(row=r_idx, column=1, value=currency)
                ws.cell(row=r_idx, column=2, value=opening_total)
                ws.cell(row=r_idx, column=3, value=closing_total)

                # Add KES conversion values if available
                col_offset = 3
                if 'opening_balance_kes' in currency_summary.columns:
                    opening_kes = row_dict.get("opening_balance_kes", 0)
                    closing_kes = row_dict.get("closing_balance_kes", 0)
                    
                    ws.cell(row=r_idx, column=4, value=opening_kes)
                    ws.cell(row=r_idx, column=5, value=closing_kes)
                    col_offset = 5

                # Apply formatting and borders
                for c in range(1, col_offset + 1):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.border = thin_border
                    if c > 1:  # All columns except currency
                        if c >= 4:  # KES conversion columns
                            cell.number_format = formats['kes_currency']
                        else:  # Original currency columns
                            cell.number_format = formats['number']
                        cell.alignment = Alignment(horizontal="right")

                # Highlight grand total row
                if currency == "GRAND TOTAL":
                    for c in range(1, col_offset + 1):
                        cell = ws.cell(row=r_idx, column=c)
                        cell.font = formats['grand_total']
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # --- Auto column width ---
            for i, col_cells in enumerate(ws.columns, start=1):
                max_length = 0
                col_letter = get_column_letter(i)
                for cell in col_cells:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        continue
                ws.column_dimensions[col_letter].width = max_length + 3

            # --- Save to in-memory stream ---
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # --- Streamlit download button ---
            st.download_button(
                label="⬇️ Download Cash Summary Excel Report",
                data=excel_buffer,
                file_name=f"Cash_Summary_{report_date.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            # --- Display summary in Streamlit ---
            st.success("✅ Cash summary report generated successfully!")
            
            # Show the currency summary table
            st.subheader("Currency Summary")
            display_summary = currency_summary.copy()
            if 'opening_balance_kes' in display_summary.columns:
                # Format for display
                for col in ['opening_balance', 'closing_balance', 'opening_balance_kes', 'closing_balance_kes']:
                    if col in display_summary.columns:
                        display_summary[col] = display_summary[col].apply(lambda x: f"{x:,.2f}")
            
            st.dataframe(display_summary, use_container_width=True)

            # Display Bank Consolidated Summary
            st.subheader("Bank Consolidated Summary (KES)")
            display_bank_summary = bank_summary_kes.copy()
            display_bank_summary["change_KES"] = display_bank_summary["closing_balance_KES"] - display_bank_summary["opening_balance_KES"]
            display_bank_summary["growth_%"] = (
                (display_bank_summary["change_KES"] / display_bank_summary["opening_balance_KES"].replace(0, float("nan"))) * 100
            )
            
            # Format for display
            for col in ['opening_balance_KES', 'closing_balance_KES', 'change_KES']:
                display_bank_summary[col] = display_bank_summary[col].apply(lambda x: f"KSh{x:,.2f}")
            display_bank_summary["growth_%"] = display_bank_summary["growth_%"].apply(lambda x: f"{x:.2f}%")
            
            st.dataframe(display_bank_summary, use_container_width=True)

# Example usage in your Streamlit app:
# generate_cash_summary_report(per_bank_df)

st.set_page_config(page_title="Finance(FX) Reconciliation Dashboard", layout="wide")

# --- Constants and Global Mappings ---
DATE_FORMATS = [
    '%Y-%m-%d', '%Y/%m/%d', '%d.%m.%Y', '%Y.%m.%d', '%d/%m/%Y',
    '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%d.%m.%Y %H:%M:%S',
    '%Y.%m.%d %H:%M:%S', '%d/%m/%Y %H:%M:%S'
]

PREDEFINED_BANK_CURRENCY_OPTIONS = [
    "Absa KES", "Absa USD", "Absa EUR", "Absa GBP", "ABSA KES-SPECIAL", "ABSA USD-SPECIAL", "ABSA EUR-SPECIAL", "ABSA GBP-SPECIAL", "ABSA Bank USD-DCD",
    "CBK KES", "CBK USD", "CBK EUR", "CBK GBP", "CBK UGX", "CBK TZS", "CBK RWF", "CBK ZAR", "CBK CNY",
    "Equity KES", "Equity USD", "Equity EUR", "Equity GBP",
    "I&M KES", "I&M USD", "I&M EUR", "I&M GBP",
    "KCB KES", "KCB USD", "KCB EUR", "KCB GBP", "KCB GBP - 1343013054",
    "Kingdom KES", "Kingdom USD", "Kingdom EUR", "Kingdom GBP",
    "NCBA KES", "NCBA USD", "NCBA EUR", "NCBA GBP",
    "SBM KES", "SBM USD", "SBM EUR", "SBM GBP",
    "UBA KES", "UBA USD", "UBA EUR", "UBA GBP",
    "BAAS Temporary KES", "BAAS Temporary USD", "BAAS Temporary EUR", "BAAS Temporary GBP",
    "FX Temporary KES", "FX Temporary USD", "FX Temporary EUR", "FX Temporary GBP",
    "Other Temporary KES", "Other Temporary USD", "Other Temporary EUR", "Other Temporary GBP",
    "Unclaimed Funds KES", "Unclaimed Funds USD", "Unclaimed Funds EUR", "Unclaimed Funds GBP",
    "Yeepay KES", "Yeepay USD", "Yeepay EUR", "Yeepay GBP", "Yeepay CNY"
]

FX_EXPECTED_COLUMNS = {
    'Amount': 'Amount', 'Operation': 'Operation', 'Completed At': 'Completed At',
    'Intermediary Account': 'Intermediary Account', 'Currency': 'Currency', 'Status': 'Status'
}

BANK_EXPECTED_COLUMNS = {
    'Date': ['Date', 'Transaction Date', 'Value Date', 'Value date'],
    'Credit': ['Credit', 'Credit Amount', 'Money In', 'Deposit', 'Credit amount'],
    'Debit': ['Debit', 'Debit Amount', 'Money Out', 'Withdrawal', 'Debit amount'],
    'Description': ['Description', 'Narrative', 'Transaction Details', 'Customer reference', 'Transaction Remarks:', 'Transaction Details', 'TransactionDetails', 'Transaction\nDetails'],
    'Running Balances': ['Running Balances', 'Running Balance', 'Running Balance (KES)', 'Running Balance (USD)', 'Running Balance (EUR)', 'Running Balance (GBP)', 'RUNNING BALANCES', 'RUNNING BALANCE', 'RUNNING BALANCE (KES)', 'RUNNING BALANCE (USD)', 'RUNNING BALANCE (EUR)', 'RUNNING BALANCE (GBP)']
}

# --- Helper Functions ---
def parse_date(date_str_raw):
    """Parses a date string into a datetime object using predefined formats."""
    if pd.isna(date_str_raw) or date_str_raw == pd.NaT: return None
    if isinstance(date_str_raw, datetime): return date_str_raw
    if not isinstance(date_str_raw, str): date_str_raw = str(date_str_raw)
    date_str = date_str_raw.partition(" ")[0].strip() if " " in date_str_raw.strip() else date_str_raw.strip()
    for fmt in DATE_FORMATS:
        try: return datetime.strptime(date_str, fmt)
        except ValueError: continue
    return None

def safe_float(x):
    """Safely converts a value to a float, handling commas, non-numeric inputs, and ensuring consistency."""
    if pd.isna(x) or x is None: return None
    try:
        cleaned_x = str(x).replace(',', '').strip()
        return float(cleaned_x)
    except (ValueError, TypeError): return None

def process_uploaded_file(uploaded_file, sheet_name=None):
    """Reads an uploaded file (CSV or Excel) into a DataFrame."""
    uploaded_file.seek(0)
    if uploaded_file.name.endswith('.csv'):
        encodings = ['utf-8', 'utf-8-sig', 'latin1', 'ISO-8859-1', 'windows-1252']
        for enc in encodings:
            try:
                df = pd.read_csv(uploaded_file, encoding=enc)
                return df
            except Exception: continue
        st.error(f"Failed to decode CSV file '{uploaded_file.name}' using common encodings.")
        return pd.DataFrame()
    elif uploaded_file.name.endswith(('.xlsx', '.xls')):
        try:
            # Handle multiple sheet selection
            if isinstance(sheet_name, list):
                # Read multiple sheets and return as dictionary
                dfs = pd.read_excel(uploaded_file, sheet_name=sheet_name)
                return dfs
            else:
                # Single sheet
                df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
                return df
        except Exception as e:
            st.error(f"Error reading Excel file '{uploaded_file.name}': {e}")
            return pd.DataFrame() if not isinstance(sheet_name, list) else {}
    else:
        st.error("Unsupported file type. Please upload a CSV or Excel file.")
        return pd.DataFrame()

def get_excel_sheet_names(uploaded_file):
    """Returns sheet names for an Excel file."""
    uploaded_file.seek(0)
    try:
        excel_file = pd.ExcelFile(uploaded_file)
        return excel_file.sheet_names
    except Exception as e:
        st.error(f"Error getting Excel sheet names: {e}")
        return []

# --- Session State Initialization ---
if 'df_matched_interfund' not in st.session_state:
    st.session_state.df_matched_interfund = pd.DataFrame()
    
if 'df_unmatched_interfund' not in st.session_state:
    st.session_state.df_unmatched_interfund = pd.DataFrame()
if 'df_matched_adjustments_local' not in st.session_state: st.session_state.df_matched_adjustments_local = pd.DataFrame()
if 'df_matched_adjustments_foreign' not in st.session_state: st.session_state.df_matched_adjustments_foreign = pd.DataFrame()
if 'df_unmatched_adjustments_local' not in st.session_state: st.session_state.df_unmatched_adjustments_local = pd.DataFrame()
if 'df_unmatched_adjustments_foreign' not in st.session_state: st.session_state.df_unmatched_adjustments_foreign = pd.DataFrame()
if 'df_unmatched_bank_recon' not in st.session_state: st.session_state.df_unmatched_bank_recon = pd.DataFrame()
if 'df_matched_counterparty' not in st.session_state: st.session_state.df_matched_counterparty = pd.DataFrame()
if 'df_matched_choice' not in st.session_state: st.session_state.df_matched_choice = pd.DataFrame()
if 'df_unmatched_counterparty' not in st.session_state: st.session_state.df_unmatched_counterparty = pd.DataFrame()
if 'df_unmatched_choice' not in st.session_state: st.session_state.df_unmatched_choice = pd.DataFrame()
if 'df_unmatched_bank_trade' not in st.session_state: st.session_state.df_unmatched_bank_trade = pd.DataFrame()
if 'df_unmatched_bank_records' not in st.session_state: st.session_state.df_unmatched_bank_records = pd.DataFrame()
if 'debug_mode' not in st.session_state: st.session_state.debug_mode = False
if 'bank_dfs' not in st.session_state: st.session_state.bank_dfs = {}
if 'bank_uploaded_file_objs' not in st.session_state: st.session_state.bank_uploaded_file_objs = []
if 'raw_bank_data_previews' not in st.session_state: st.session_state.raw_bank_data_previews = {}
if 'merged_bank_statement' not in st.session_state: st.session_state.merged_bank_statement = pd.DataFrame()
if "cached_bank_files" not in st.session_state: st.session_state.cached_bank_files = {}
page_selection = st.sidebar.radio("Go to", ["Bank Statement Management", "Adjacements Reconciliation", "FX Trade Reconciliation", "Intermediary Reconciliation", "Interfund Reconciliation", "Business FX Reconciliation", "Cross-Match Analysis"])


# Add the detection function from your reference code
def detect_column_type(series):
    """
    Detect if a pandas series contains datetime objects or strings
    """
    # Check if already datetime
    if pd.api.types.is_datetime64_any_dtype(series):
        return 'datetime'
    
    # Check if numeric (not date)
    if pd.api.types.is_numeric_dtype(series):
        return 'numeric'
    
    # Try to convert sample to datetime to check if it contains dates
    sample_size = min(100, len(series))
    sample = series.head(sample_size).dropna()
    
    if len(sample) == 0:
        return 'unknown'
    
    # Try parsing as datetime
    try:
        # Try multiple common date formats
        test_parsed = pd.to_datetime(sample, errors='coerce', format='%m/%d/%Y')
        success_rate = (test_parsed.notna().sum() / len(sample)) * 100
        
        if success_rate > 80:  # If >80% successfully parsed as dates
            return 'date_string'
        else:
            return 'general_string'
    except:
        return 'general_string'

# --- Main App Logic ---
if page_selection == "Bank Statement Management":
    st.title("Bank Statement Management")
    st.markdown("Upload and configure your bank statements here. These statements will then be available for all reconciliation modules.")

    uploaded_files = st.file_uploader("Upload Bank Statement(s) (CSV/Excel)", type=["csv", "xlsx"], accept_multiple_files=True, key="bank_uploader_main")
    
    if uploaded_files:
        for file in uploaded_files:
            if file.name not in st.session_state.cached_bank_files:
                file_bytes = file.read()
                file_type = file.type
                st.session_state.cached_bank_files[file.name] = {"content": file_bytes, "type": file_type}

    files_to_delete = []

    if st.session_state.cached_bank_files:
        st.markdown("### Uploaded Bank Statements:")
        for file_name, file_data in st.session_state.cached_bank_files.items():
            file_key = file_name.lower().replace('.', '_')

            with st.expander(f"🗂️ {file_name}", expanded=True):
                col1, col2 = st.columns([8, 2])
                with col1: st.markdown(f"**File Name:** `{file_name}`")
                with col2:
                    if st.button("❌ Remove", key=f"remove_{file_name}"):
                        files_to_delete.append(file_name)
                        continue

                if file_key not in st.session_state.raw_bank_data_previews:
                    fake_file = BytesIO(file_data["content"])
                    fake_file.name = file_name

                    if file_name.endswith('.xlsx'):
                        sheet_names = get_excel_sheet_names(fake_file)
                        selected_sheets = [sheet_names[0]] if sheet_names else []
                        # Read all selected sheets
                        if selected_sheets:
                            dfs = process_uploaded_file(fake_file, sheet_name=selected_sheets)
                            # Store as dictionary of DataFrames
                            df_dict = dfs if isinstance(dfs, dict) else {selected_sheets[0]: dfs}
                        else:
                            df_dict = {}
                    else:
                        sheet_names = []
                        selected_sheets = []
                        df_dict = {"CSV": process_uploaded_file(fake_file)}

                    # Initialize standardized names for each sheet
                    standardized_names = {}
                    for sheet_name in df_dict.keys():
                        standardized_names[sheet_name] = ""

                    st.session_state.raw_bank_data_previews[file_key] = {
                        'file_obj': fake_file, 
                        'df_raw_dict': df_dict, 
                        'sheet_names': sheet_names,
                        'selected_sheets': selected_sheets,
                        'column_mappings': {}, 
                        'standardized_names': standardized_names
                    }

                data = st.session_state.raw_bank_data_previews[file_key]
                df_bank_raw_dict = data['df_raw_dict']

                if file_name.endswith('.xlsx') and data['sheet_names']:
                    current_sheets = st.multiselect(
                        f"Select Sheets for {file_name}:", 
                        data['sheet_names'],
                        default=data['selected_sheets'],
                        key=f"bank_sheet_selector_{file_key}"
                    )
                    
                    if set(current_sheets) != set(data['selected_sheets']):
                        data['selected_sheets'] = current_sheets
                        if current_sheets:
                            fake_file = BytesIO(file_data["content"])
                            fake_file.name = file_name
                            dfs = process_uploaded_file(fake_file, sheet_name=current_sheets)
                            df_bank_raw_dict = dfs if isinstance(dfs, dict) else {current_sheets[0]: dfs}
                            for sheet_name, df in df_bank_raw_dict.items():
                                if df is not None:
                                    df.columns = df.columns.str.strip()
                            
                            for sheet_name in df_bank_raw_dict.keys():
                                if sheet_name not in data['standardized_names']:
                                    data['standardized_names'][sheet_name] = ""
                            
                            st.info(f"Selected {len(current_sheets)} sheet(s) for {file_name}.")
                        else:
                            df_bank_raw_dict = {}
                        data['df_raw_dict'] = df_bank_raw_dict

                # Display standardized name selector for each sheet
                if df_bank_raw_dict:
                    for sheet_name, df_bank_raw in df_bank_raw_dict.items():
                        if df_bank_raw is not None and not df_bank_raw.empty:
                            st.markdown(f"---")
                            st.subheader(f"Sheet: {sheet_name}")
                            
                            # Process datetime columns for this sheet
                            if sheet_name in data['df_raw_dict']:
                                df_processed = data['df_raw_dict'][sheet_name].copy()
                                conversion_log = []
                                
                                # Process each column for datetime conversion
                                for col in df_processed.columns:
                                    # Detect column type using the reference logic
                                    col_type = detect_column_type(df_processed[col])
                                    
                                    # if col_type in ['datetime', 'date_string']:
                                    if col_type in ['datetime', 'date_string']:

                                        # Convert datetime columns
                                        try:
                                            # Store original for comparison
                                            original_sample = df_processed[col].head(3).tolist()
                                            
                                            # Convert to datetime
                                            df_processed[col] = pd.to_datetime(
                                                df_processed[col], 
                                                format='%m/%d/%Y',
                                                errors='coerce'
                                            )
                                            
                                            # Count successful conversions
                                            successful_conversions = df_processed[col].notna().sum()
                                            total_rows = len(df_processed)
                                            
                                            if successful_conversions > 0:
                                                # Format as dd/m/yyyy for display
                                                df_processed[col] = df_processed[col].dt.strftime('%d/%m/%Y')
                                                conversion_log.append(f"✅ **{col}**: Converted {successful_conversions}/{total_rows} dates")
                                                
                                                # Show before/after sample
                                                with st.expander(f"Show conversion samples for '{col}'"):
                                                    st.write("**Before:**", original_sample)
                                                    st.write("**After:**", df_processed[col].head(3).tolist())
                                            else:
                                                conversion_log.append(f"❌ **{col}**: No dates successfully converted")
                                                # Revert to original if conversion failed
                                                df_processed[col] = data['df_raw_dict'][sheet_name][col]
                                                
                                        except Exception as e:
                                            conversion_log.append(f"❌ **{col}**: Error - {str(e)}")
                                            # Revert to original on error
                                            df_processed[col] = data['df_raw_dict'][sheet_name][col]
                                    elif col_type == 'general_string':
                                        conversion_log.append(f"ℹ️ **{col}**: General text (no conversion needed)")
                                    elif col_type == 'numeric':
                                        conversion_log.append(f"ℹ️ **{col}**: Numeric data (no conversion needed)")
                                
                                # Update the dataframe in the dictionary
                                data['df_raw_dict'][sheet_name] = df_processed
                                
                                # Show conversion summary
                                if conversion_log:
                                    with st.expander("Date Conversion Summary"):
                                        for log_entry in conversion_log:
                                            st.write(log_entry)
                            
                            # Standardized name selector for this specific sheet
                            selected_standardized_name = st.selectbox(
                                f"Select Standardized Name for '{sheet_name}':", 
                                options=[""] + PREDEFINED_BANK_CURRENCY_OPTIONS,
                                index=PREDEFINED_BANK_CURRENCY_OPTIONS.index(data['standardized_names'].get(sheet_name, "")) + 1 
                                if data['standardized_names'].get(sheet_name, "") in PREDEFINED_BANK_CURRENCY_OPTIONS else 0,
                                key=f"standardized_name_selector_{file_key}_{sheet_name}"
                            )
                            data['standardized_names'][sheet_name] = selected_standardized_name

                            st.write(f"**Preview - {sheet_name}:**")
                            # Use the processed dataframe for display
                            display_df = data['df_raw_dict'][sheet_name]
                            st.dataframe(display_df.head())

                            available_columns = display_df.columns.tolist()
                            available_columns.insert(0, "")
                            # Initialize column mappings for this sheet if it doesn't exist
                            if sheet_name not in data['column_mappings']:
                                data['column_mappings'][sheet_name] = {}
                            current_mappings = data['column_mappings'][sheet_name]

                            st.write(f"**Column Mapping - {sheet_name}:**")
                            col_map_cols = st.columns(2)
                            for expected_col, default_val_list in BANK_EXPECTED_COLUMNS.items():
                                initial_selection = current_mappings.get(expected_col)
                                if not initial_selection:
                                    for default_val in default_val_list:
                                        if default_val.strip() in [col.strip() for col in display_df.columns]:
                                            initial_selection = default_val
                                            break
                                
                                with col_map_cols[0]: st.markdown(f"**{expected_col}**")
                                with col_map_cols[1]:
                                    mapped_col = st.selectbox(
                                        f"Map '{expected_col}' to ({sheet_name}):", options=available_columns,
                                        index=available_columns.index(initial_selection) if initial_selection and initial_selection in available_columns else 0,
                                        key=f"bank_map_{file_key}_{sheet_name}_{expected_col}",
                                        label_visibility="collapsed"
                                    )
                                    data['column_mappings'][sheet_name][expected_col] = mapped_col if mapped_col else None
                        else:
                            st.warning(f"No data loaded for sheet '{sheet_name}' in {file_name}.")
                else:
                    st.error(f"Could not load data from {file_name}.")
    
    for file_name in files_to_delete:
        st.session_state.cached_bank_files.pop(file_name, None)
        file_key = file_name.lower().replace('.', '_')
        st.session_state.raw_bank_data_previews.pop(file_key, None)
        st.success(f"File '{file_name}' and its data have been removed.")

    st.session_state.bank_dfs = {}
    st.session_state.merged_bank_statement = pd.DataFrame()




# # Add the detection function from your reference code
# def detect_column_type(series):
#     """
#     Detect if a pandas series contains datetime objects or strings
#     """
#     # Check if already datetime
#     if pd.api.types.is_datetime64_any_dtype(series):
#         return 'datetime'
    
#     # Check if numeric (not date)
#     if pd.api.types.is_numeric_dtype(series):
#         return 'numeric'
    
#     # Try to convert sample to datetime to check if it contains dates
#     sample_size = min(100, len(series))
#     sample = series.head(sample_size).dropna()
    
#     if len(sample) == 0:
#         return 'unknown'
    
#     # Try parsing as datetime
#     try:
#         # Try multiple common date formats
#         test_parsed = pd.to_datetime(sample, errors='coerce')
#         success_rate = (test_parsed.notna().sum() / len(sample)) * 100
        
#         if success_rate > 80:  # If >80% successfully parsed as dates
#             return 'date_string'
#         else:
#             return 'general_string'
#     except:
#         return 'general_string'
    


# # --- Main App Logic ---
# if page_selection == "Bank Statement Management":
#     st.title("Bank Statement Management")
#     st.markdown("Upload and configure your bank statements here. These statements will then be available for all reconciliation modules.")

#     uploaded_files = st.file_uploader("Upload Bank Statement(s) (CSV/Excel)", type=["csv", "xlsx"], accept_multiple_files=True, key="bank_uploader_main")
    
#     if uploaded_files:
#         for file in uploaded_files:
#             if file.name not in st.session_state.cached_bank_files:
#                 file_bytes = file.read()
#                 file_type = file.type
#                 st.session_state.cached_bank_files[file.name] = {"content": file_bytes, "type": file_type}

#     files_to_delete = []

#     if st.session_state.cached_bank_files:
#         st.markdown("### Uploaded Bank Statements:")
#         for file_name, file_data in st.session_state.cached_bank_files.items():
#             file_key = file_name.lower().replace('.', '_')

#             with st.expander(f"🗂️ {file_name}", expanded=True):
#                 col1, col2 = st.columns([8, 2])
#                 with col1: st.markdown(f"**File Name:** `{file_name}`")
#                 with col2:
#                     if st.button("❌ Remove", key=f"remove_{file_name}"):
#                         files_to_delete.append(file_name)
#                         continue

#                 if file_key not in st.session_state.raw_bank_data_previews:
#                     fake_file = BytesIO(file_data["content"])
#                     fake_file.name = file_name

#                     if file_name.endswith('.xlsx'):
#                         sheet_names = get_excel_sheet_names(fake_file)
#                         selected_sheets = [sheet_names[0]] if sheet_names else []
#                         # Read all selected sheets
#                         if selected_sheets:
#                             dfs = process_uploaded_file(fake_file, sheet_name=selected_sheets)
#                             # Store as dictionary of DataFrames
#                             df_dict = dfs if isinstance(dfs, dict) else {selected_sheets[0]: dfs}
#                         else:
#                             df_dict = {}
#                     else:
#                         sheet_names = []
#                         selected_sheets = []
#                         df_dict = {"CSV": process_uploaded_file(fake_file)}

#                     # Initialize standardized names for each sheet
#                     standardized_names = {}
#                     for sheet_name in df_dict.keys():
#                         standardized_names[sheet_name] = ""

#                     st.session_state.raw_bank_data_previews[file_key] = {
#                         'file_obj': fake_file, 
#                         'df_raw_dict': df_dict, 
#                         'sheet_names': sheet_names,
#                         'selected_sheets': selected_sheets,
#                         'column_mappings': {}, 
#                         'standardized_names': standardized_names
#                     }

#                 data = st.session_state.raw_bank_data_previews[file_key]
#                 df_bank_raw_dict = data['df_raw_dict']

#                 if file_name.endswith('.xlsx') and data['sheet_names']:
#                     current_sheets = st.multiselect(
#                         f"Select Sheets for {file_name}:", 
#                         data['sheet_names'],
#                         default=data['selected_sheets'],
#                         key=f"bank_sheet_selector_{file_key}"
#                     )
                    
#                     if set(current_sheets) != set(data['selected_sheets']):
#                         data['selected_sheets'] = current_sheets
#                         if current_sheets:
#                             fake_file = BytesIO(file_data["content"])
#                             fake_file.name = file_name
#                             dfs = process_uploaded_file(fake_file, sheet_name=current_sheets)
#                             df_bank_raw_dict = dfs if isinstance(dfs, dict) else {current_sheets[0]: dfs}
#                             for sheet_name, df in df_bank_raw_dict.items():
#                                 if df is not None:
#                                     df.columns = df.columns.str.strip()
                            
#                             for sheet_name in df_bank_raw_dict.keys():
#                                 if sheet_name not in data['standardized_names']:
#                                     data['standardized_names'][sheet_name] = ""
                            
#                             st.info(f"Selected {len(current_sheets)} sheet(s) for {file_name}.")
#                         else:
#                             df_bank_raw_dict = {}
#                         data['df_raw_dict'] = df_bank_raw_dict

#                 # Display standardized name selector for each sheet
#                 if df_bank_raw_dict:
#                     for sheet_name, df_bank_raw in df_bank_raw_dict.items():
#                         if df_bank_raw is not None and not df_bank_raw.empty:
#                             st.markdown(f"---")
#                             st.subheader(f"Sheet: {sheet_name}")
                            
#                             # Process datetime columns for this sheet
#                             if sheet_name in data['df_raw_dict']:
#                                 df_processed = data['df_raw_dict'][sheet_name].copy()
#                                 conversion_log = []
                                
#                                 # Process each column for datetime conversion
#                                 for col in df_processed.columns:
#                                     # Detect column type using the reference logic
#                                     col_type = detect_column_type(df_processed[col])
                                    
#                                     if col_type in ['datetime', 'date_string']:
#                                         # Convert datetime columns
#                                         try:
#                                             # Store original for comparison
#                                             original_sample = df_processed[col].head(3).tolist()
                                            
#                                             # Convert to datetime
#                                             df_processed[col] = pd.to_datetime(
#                                                 df_processed[col], 
#                                                 errors='coerce'
#                                             )
                                            
#                                             # Count successful conversions
#                                             successful_conversions = df_processed[col].notna().sum()
#                                             total_rows = len(df_processed)
                                            
#                                             if successful_conversions > 0:
#                                                 # Format as mm/dd/yyyy for display (month/day/year)
#                                                 df_processed[col] = df_processed[col].dt.strftime('%m/%d/%Y')
#                                                 conversion_log.append(f"✅ **{col}**: Converted {successful_conversions}/{total_rows} dates to mm/dd/yyyy")
                                                
#                                                 # Show before/after sample
#                                                 with st.expander(f"Show conversion samples for '{col}'"):
#                                                     st.write("**Before:**", original_sample)
#                                                     st.write("**After:**", df_processed[col].head(3).tolist())
#                                             else:
#                                                 conversion_log.append(f"❌ **{col}**: No dates successfully converted")
#                                                 # Revert to original if conversion failed
#                                                 df_processed[col] = data['df_raw_dict'][sheet_name][col]
                                                
#                                         except Exception as e:
#                                             conversion_log.append(f"❌ **{col}**: Error - {str(e)}")
#                                             # Revert to original on error
#                                             df_processed[col] = data['df_raw_dict'][sheet_name][col]
#                                     elif col_type == 'general_string':
#                                         conversion_log.append(f"ℹ️ **{col}**: General text (no conversion needed)")
#                                     elif col_type == 'numeric':
#                                         conversion_log.append(f"ℹ️ **{col}**: Numeric data (no conversion needed)")
                                
#                                 # Update the dataframe in the dictionary
#                                 data['df_raw_dict'][sheet_name] = df_processed
                                
#                                 # Show conversion summary
#                                 if conversion_log:
#                                     with st.expander("Date Conversion Summary"):
#                                         for log_entry in conversion_log:
#                                             st.write(log_entry)
                            
#                             # Standardized name selector for this specific sheet
#                             selected_standardized_name = st.selectbox(
#                                 f"Select Standardized Name for '{sheet_name}':", 
#                                 options=[""] + PREDEFINED_BANK_CURRENCY_OPTIONS,
#                                 index=PREDEFINED_BANK_CURRENCY_OPTIONS.index(data['standardized_names'].get(sheet_name, "")) + 1 
#                                 if data['standardized_names'].get(sheet_name, "") in PREDEFINED_BANK_CURRENCY_OPTIONS else 0,
#                                 key=f"standardized_name_selector_{file_key}_{sheet_name}"
#                             )
#                             data['standardized_names'][sheet_name] = selected_standardized_name

#                             st.write(f"**Preview - {sheet_name}:**")
#                             # Use the processed dataframe for display
#                             display_df = data['df_raw_dict'][sheet_name]
#                             st.dataframe(display_df.head())

#                             available_columns = display_df.columns.tolist()
#                             available_columns.insert(0, "")
#                             # Initialize column mappings for this sheet if it doesn't exist
#                             if sheet_name not in data['column_mappings']:
#                                 data['column_mappings'][sheet_name] = {}
#                             current_mappings = data['column_mappings'][sheet_name]

#                             st.write(f"**Column Mapping - {sheet_name}:**")
#                             col_map_cols = st.columns(2)
#                             for expected_col, default_val_list in BANK_EXPECTED_COLUMNS.items():
#                                 initial_selection = current_mappings.get(expected_col)
#                                 if not initial_selection:
#                                     for default_val in default_val_list:
#                                         if default_val.strip() in [col.strip() for col in display_df.columns]:
#                                             initial_selection = default_val
#                                             break
                                
#                                 with col_map_cols[0]: st.markdown(f"**{expected_col}**")
#                                 with col_map_cols[1]:
#                                     mapped_col = st.selectbox(
#                                         f"Map '{expected_col}' to ({sheet_name}):", options=available_columns,
#                                         index=available_columns.index(initial_selection) if initial_selection and initial_selection in available_columns else 0,
#                                         key=f"bank_map_{file_key}_{sheet_name}_{expected_col}",
#                                         label_visibility="collapsed"
#                                     )
#                                     data['column_mappings'][sheet_name][expected_col] = mapped_col if mapped_col else None
#                         else:
#                             st.warning(f"No data loaded for sheet '{sheet_name}' in {file_name}.")
#                 else:
#                     st.error(f"Could not load data from {file_name}.")
    
#     for file_name in files_to_delete:
#         st.session_state.cached_bank_files.pop(file_name, None)
#         file_key = file_name.lower().replace('.', '_')
#         st.session_state.raw_bank_data_previews.pop(file_key, None)
#         st.success(f"File '{file_name}' and its data have been removed.")

#     st.session_state.bank_dfs = {}
#     st.session_state.merged_bank_statement = pd.DataFrame()



#------------------------------------------------------
    if st.button("Process All Bank Statements", key="process_all_bank_btn_main"):
        st.session_state.bank_dfs = {}
        all_success = True
        dfs_to_concat = []
        st.session_state.running_balances_col = None

        for file_key, data in st.session_state.raw_bank_data_previews.items():
            st.info(f"Processing '{data['file_obj'].name}'...")

            # Process each sheet separately
            sheet_dfs = []
            for sheet_name, df_raw in data['df_raw_dict'].items():
                if df_raw is None or df_raw.empty:
                    st.warning(f"Skipping empty sheet '{sheet_name}' in '{data['file_obj'].name}'")
                    continue

                # Check standardized name for this specific sheet - FIXED
                sheet_standardized_name = data['standardized_names'].get(sheet_name, "")
                if not sheet_standardized_name:
                    st.error(f"Missing standardized name for sheet '{sheet_name}' in '{data['file_obj'].name}'")
                    all_success = False
                    continue

                # Check for duplicate standardized names across all sheets
                if sheet_standardized_name in st.session_state.bank_dfs:
                    st.error(f"Duplicate standardized name '{sheet_standardized_name}' detected for sheet '{sheet_name}'. Please choose a unique name for each sheet.")
                    all_success = False
                    continue

                df_to_process = df_raw.copy()
                
                # Get column mappings for this specific sheet
                sheet_mappings = data['column_mappings'].get(sheet_name, {})
                
                renamed_cols = {}
                for expected_col, mapped_col in sheet_mappings.items():
                    if mapped_col and mapped_col in df_to_process.columns:
                        renamed_cols[mapped_col] = expected_col
                
                if renamed_cols:
                    df_to_process.rename(columns=renamed_cols, inplace=True)
                df_to_process.columns = df_to_process.columns.str.strip()
                
                # --- Advanced Data Validation ---
                required_cols = ['Date', 'Credit', 'Debit', 'Running Balances']
                missing_cols = [col for col in required_cols if col not in df_to_process.columns]
                if missing_cols:
                    st.error(f"Validation failed for sheet '{sheet_name}' in '{data['file_obj'].name}'. Missing columns: {', '.join(missing_cols)}.")
                    all_success = False
                    continue

                # Process data
                df_to_process['Date'] = df_to_process['Date'].apply(parse_date)
                invalid_dates_mask = df_to_process['Date'].isna()
                if invalid_dates_mask.any():
                    num_errors = invalid_dates_mask.sum()
                    st.warning(f"Warning in sheet '{sheet_name}' of '{data['file_obj'].name}': {num_errors} invalid dates found. These rows will be dropped.")
                    df_to_process = df_to_process[~invalid_dates_mask].copy()

                df_to_process['Credit'] = df_to_process['Credit'].apply(safe_float)
                df_to_process['Debit'] = df_to_process['Debit'].apply(safe_float)
                df_to_process['Running Balances'] = df_to_process['Running Balances'].apply(safe_float)

                df_to_process["Matched"] = False
                df_to_process['Bank'] = sheet_standardized_name  # Use sheet-specific standardized name
                df_to_process['Source_Sheet'] = sheet_name  # Track which sheet this data came from
                df_to_process['Source_File'] = data['file_obj'].name  # Track which file this data came from
                
                sheet_dfs.append(df_to_process)
                st.success(f"Processed: {data['file_obj'].name} - Sheet '{sheet_name}' as '{sheet_standardized_name}'")

            # Combine all sheets from this file
            if sheet_dfs:
                # Store each sheet's data in bank_dfs
                for df in sheet_dfs:
                    bank_name = df['Bank'].iloc[0] if not df.empty else None
                    if bank_name:
                        st.session_state.bank_dfs[bank_name] = df
                dfs_to_concat.extend(sheet_dfs)
            else:
                st.error(f"No valid sheets found in '{data['file_obj'].name}'")
                all_success = False

        if all_success and dfs_to_concat:
            st.session_state.merged_bank_statement = pd.concat(dfs_to_concat, ignore_index=True)
            st.write("✅ All bank statements processed and merged.")
            
            if not st.session_state.merged_bank_statement.empty:
                df_bal = st.session_state.merged_bank_statement.copy()
                rb_col = 'Running Balances' # This column is now standardized in all dataframes
                
                df_bal.rename(columns={'Date': 'date', 'Debit': 'debit', 'Credit': 'credit', 'Bank': 'bank'}, inplace=True)
                df_bal["currency"] = df_bal["bank"].apply(lambda x: str(x).split()[-1].upper())
                df_bal = df_bal.sort_values(by=['bank', 'date'])
                
                per_bank_rows = []
                for bank_name, df_bank in df_bal.groupby("bank"):
                    df_bank = df_bank.sort_values("date").reset_index(drop=True)
                    first_row = df_bank.iloc[0]
                    last_row = df_bank.iloc[-1]
                    currency = str(bank_name).split()[-1].upper()
                    running_balance_first = first_row[rb_col] if pd.notna(first_row[rb_col]) else 0
                    debit_first = first_row["debit"] if pd.notna(first_row["debit"]) else 0
                    credit_first = first_row["credit"] if pd.notna(first_row["credit"]) else 0

                    opening_balance = running_balance_first - credit_first + debit_first
                    closing_balance = last_row[rb_col] if pd.notna(last_row[rb_col]) else 0

                    per_bank_rows.append({"Bank": bank_name, "Currency": currency, "Opening Balance": round(opening_balance, 2), "Closing Balance": round(closing_balance, 2)})

                per_bank_df = pd.DataFrame(per_bank_rows).sort_values(by=["Currency", "Bank"]).reset_index(drop=True)
                st.subheader("Per-Bank Opening & Closing Balances")
                st.dataframe(per_bank_df)
                csv_per_bank = per_bank_df.to_csv(index=False).encode("utf-8")
                st.download_button(label="⬇️ Download Per-Bank Balances CSV", data=csv_per_bank, file_name="per_bank_balances.csv", mime="text/csv")
                
                currency_summary = (per_bank_df.groupby("Currency").agg({"Opening Balance": "sum", "Closing Balance": "sum"}).round(2).reset_index().sort_values(by="Currency").reset_index(drop=True))
                st.subheader("Opening & Closing Balance Summary by Currency")
                st.dataframe(currency_summary)
                csv_summary = currency_summary.to_csv(index=False).encode("utf-8")
                st.download_button(label="⬇️ Download Currency Summary CSV", data=csv_summary, file_name="currency_balance_summary.csv", mime="text/csv")

                st.markdown("---")
                st.subheader("Monthly Transaction Volume")
                df_chart = st.session_state.merged_bank_statement.copy()
                df_chart['YearMonth'] = pd.to_datetime(df_chart['Date']).dt.to_period('M').astype(str)
                df_chart['Credit'] = pd.to_numeric(df_chart['Credit'], errors='coerce').fillna(0)
                df_chart['Debit'] = pd.to_numeric(df_chart['Debit'], errors='coerce').fillna(0)
                monthly_volume = df_chart.groupby(['Bank', 'YearMonth']).agg(
                    Total_Credit=('Credit', 'sum'),
                    Total_Debit=('Debit', 'sum')
                ).reset_index()
                st.bar_chart(monthly_volume, x='YearMonth', y=['Total_Credit', 'Total_Debit'], color=['#008000', '#FF0000'])
            
        elif all_success and not dfs_to_concat: 
            st.info("⚠️ No valid files processed.")
        else: 
            st.warning("⚠️ Some files could not be processed. See messages above.")
        st.markdown("---")
        st.header("Merged Bank Statement for Display and Download")
        if not st.session_state.get("merged_bank_statement", pd.DataFrame()).empty:
            st.write("### Combined Merged Statement:")
            st.dataframe(st.session_state.merged_bank_statement)
            csv = st.session_state.merged_bank_statement.to_csv(index=False).encode("utf-8")
            st.download_button(label="⬇️ Download Merged Bank Statement as CSV", data=csv, file_name="merged_bank_statement.csv", mime="text/csv")
        else: 
            st.info("No merged bank statement available yet.")
        per_bank_df = per_bank_df if 'per_bank_df' in locals() else pd.DataFrame()

        # --- Ensure data exists ---
        if not per_bank_df.empty:
            generate_cash_summary_report(per_bank_df)
            
            # --- 🧭 Identify possible columns automatically ---
            currency_col_candidates = ["Currency", "currency", "CURRENCY", "Curr", "curr", "Ccy"]
            opening_col_candidates = ["Opening Balance", "opening_balance", "Open Bal", "Opening", "Opening_Balance"]
            closing_col_candidates = ["Closing Balance", "closing_balance", "Close Bal", "Closing", "Closing_Balance"]

            # Find best matches
            currency_col = next((col for col in per_bank_df.columns if col in currency_col_candidates), None)
            opening_col = next((col for col in per_bank_df.columns if col in opening_col_candidates), None)
            closing_col = next((col for col in per_bank_df.columns if col in closing_col_candidates), None)

            # --- Validate ---
            if not all([currency_col, opening_col, closing_col]):
                st.error(f"❌ Missing columns in data. Found: {list(per_bank_df.columns)}")
            else:
                # --- 🧩 Normalize and combine currency variants ---
                currency_map = {
                    "KES-SPECIAL": "KES",
                    "USD-SPECIAL": "USD",
                    "EUR-SPECIAL": "EUR",
                    "GBP-SPECIAL": "GBP",
                    "USD-DCD": "USD",
                }

                per_bank_df[currency_col] = per_bank_df[currency_col].replace(currency_map)

                # --- 🧮 Recompute normalized currency summary ---
                currency_summary = (
                    per_bank_df.groupby(currency_col, as_index=False)
                    .agg({
                        opening_col: "sum",
                        closing_col: "sum"
                    })
                    .rename(columns={
                        opening_col: "opening_balance",
                        closing_col: "closing_balance",
                        currency_col: "currency"
                    })
                )

                # --- Remove any "GRAND TOTAL" rows if they exist ---
                analytics_df = currency_summary[currency_summary["currency"].str.upper() != "GRAND TOTAL"].copy()

                # --- 💱 Currency Conversion (to KES equivalent) ---
                fx_rates = {
                    "KES": 1.0,
                    "USD": 130.0,
                    "EUR": 140.0,
                    "GBP": 160.0,
                }

                analytics_df["fx_rate_to_KES"] = analytics_df["currency"].map(fx_rates).fillna(1.0)
                analytics_df["opening_balance_KES"] = analytics_df["opening_balance"] * analytics_df["fx_rate_to_KES"]
                analytics_df["closing_balance_KES"] = analytics_df["closing_balance"] * analytics_df["fx_rate_to_KES"]

                # --- 📊 Compute difference and growth % ---
                analytics_df["change"] = analytics_df["closing_balance"] - analytics_df["opening_balance"]
                analytics_df["growth_%"] = (
                    analytics_df["change"] / analytics_df["opening_balance"].replace(0, float("nan"))
                ) * 100

                # --- 🧠 Dashboard Header ---
                st.markdown("## 📈 Cash Summary Analytics Dashboard")
                st.markdown("""
                This dashboard provides visual insights into **cash holdings by currency**,  
                automatically merging related variants (e.g., `USD-DCD`, `KES-SPECIAL`)  
                and converting totals into **KES equivalents**.
                """)

                # === 1️⃣ Currency Distribution (Closing Balances in KES) ===
                fig1 = go.Figure(data=[
                    go.Pie(
                        labels=analytics_df["currency"],
                        values=analytics_df["closing_balance_KES"],
                        hole=0.4,
                        textinfo="label+percent",
                        insidetextorientation="radial"
                    )
                ])
                fig1.update_layout(
                    title_text="💰 Distribution of Cash by Currency (KES Equivalent)",
                    showlegend=True
                )
                st.plotly_chart(fig1, use_container_width=True)

                st.markdown("""
                **Insight:**  
                - Displays **KES-equivalent distribution** of all currencies.  
                - Combines related types (`KES-SPECIAL`, `USD-DCD`, etc.) into unified totals.  
                - Helps gauge **FX exposure** and **cash diversification**.
                """)

                # === 2️⃣ Opening vs. Closing Balance Comparison (KES equivalent) ===
                fig2 = go.Figure()
                fig2.add_trace(go.Bar(
                    x=analytics_df["currency"],
                    y=analytics_df["opening_balance_KES"],
                    name="Opening (KES)",
                    marker_color="royalblue"
                ))
                fig2.add_trace(go.Bar(
                    x=analytics_df["currency"],
                    y=analytics_df["closing_balance_KES"],
                    name="Closing (KES)",
                    marker_color="seagreen"
                ))

                fig2.update_layout(
                    barmode="group",
                    title="🏦 Opening vs. Closing Balances (KES Equivalent)",
                    xaxis_title="Currency",
                    yaxis_title="KES Amount",
                    legend_title="Balance Type"
                )
                st.plotly_chart(fig2, use_container_width=True)

                st.markdown("""
                **Insight:**  
                - Compares **opening vs. closing totals** by currency (converted to KES).  
                - A higher closing bar indicates inflow or accumulation; a lower one signals outflow.
                """)

                # === 3️⃣ Growth or Decline Percentage ===
                fig3 = go.Figure()
                fig3.add_trace(go.Bar(
                    x=analytics_df["currency"],
                    y=analytics_df["growth_%"],
                    text=analytics_df["growth_%"].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else ""),
                    textposition="outside",
                    marker_color=analytics_df["growth_%"].apply(lambda x: "seagreen" if x >= 0 else "crimson")
                ))
                fig3.update_layout(
                    title="📊 Growth/Decline by Currency (%)",
                    xaxis_title="Currency",
                    yaxis_title="Growth %",
                    yaxis_tickformat=".1f",
                    showlegend=False
                )
                st.plotly_chart(fig3, use_container_width=True)

                st.markdown("""
                **Insight:**  
                - Green = Growth, Red = Decline.  
                - Useful for identifying strong or weakening currency segments in your portfolio.
                """)

                # === 4️⃣ Detailed Summary Table ===
                st.markdown("### 📋 Detailed Currency Performance Summary")
                st.dataframe(
                    analytics_df[[
                        "currency",
                        "opening_balance",
                        "closing_balance",
                        "opening_balance_KES",
                        "closing_balance_KES",
                        "change",
                        "growth_%"
                    ]]
                    .style.format({
                        "opening_balance": "₤{:,.2f}".format,
                        "closing_balance": "₤{:,.2f}".format,
                        "opening_balance_KES": "KSh{:,.2f}".format,
                        "closing_balance_KES": "KSh{:,.2f}".format,
                        "change": "₤{:,.2f}".format,
                        "growth_%": "{:.2f}%".format,
                    })
                    .applymap(lambda v: "color: green" if isinstance(v, (int, float)) and v > 0 else ("color: red" if isinstance(v, (int, float)) and v < 0 else ""))
                )

                st.markdown("""
                **Interpretation:**  
                - Includes both **native** and **KES-converted** values.  
                - `Change` and `Growth %` show absolute and relative performance.  
                - Use to track liquidity and FX movement over time.
                """)

                        # === 🏦 True Bank-Level Consolidated Summary (KES Equivalent) ===
            st.markdown("## 🏦 Bank Consolidated Summary (All Currencies Combined to KES)")

            bank_col_candidates = ["Bank", "bank", "BANK"]
            bank_col = next((col for col in per_bank_df.columns if col in bank_col_candidates), None)

            if not bank_col:
                st.warning("⚠️ 'Bank' column not found for consolidated bank summary.")
            else:
                bank_summary_df = per_bank_df.copy()

                # --- Normalize currency variants ---
                currency_map = {
                    "KES-SPECIAL": "KES",
                    "USD-SPECIAL": "USD",
                    "EUR-SPECIAL": "EUR",
                    "GBP-SPECIAL": "GBP",
                    "USD-DCD": "USD",
                }
                bank_summary_df["currency"] = bank_summary_df["currency"].replace(currency_map)

                # --- Ensure consistent FX rates ---
                if "fx_rates" not in locals():
                    fx_rates = {
                        "KES": 1.0,
                        "USD": 130.0,
                        "EUR": 140.0,
                        "GBP": 160.0,
                    }

                # --- Derive true bank name (strip currency suffixes) ---
                # Examples: "NCBA USD" → "NCBA", "ABSA KES" → "ABSA"
                bank_summary_df["bank_clean"] = (
                    bank_summary_df[bank_col]
                    .astype(str)
                    .str.replace(r"\b(USD|EUR|GBP|KES|CNY|ZAR|TZS|UGX|RWF)\b", "", regex=True)
                    .str.replace(r"[-_/]+$", "", regex=True)
                    .str.strip()
                )

                # --- Convert all balances to KES equivalent ---
                bank_summary_df["fx_rate_to_KES"] = bank_summary_df["currency"].map(fx_rates).fillna(1.0)
                bank_summary_df["opening_balance_KES"] = bank_summary_df["opening_balance"] * bank_summary_df["fx_rate_to_KES"]
                bank_summary_df["closing_balance_KES"] = bank_summary_df["closing_balance"] * bank_summary_df["fx_rate_to_KES"]

                # --- Consolidate by clean bank name ---
                bank_summary_kes = (
                    bank_summary_df.groupby("bank_clean", as_index=False)[["opening_balance_KES", "closing_balance_KES"]]
                    .sum()
                    .sort_values("closing_balance_KES", ascending=False)
                )

                # --- Compute growth and change ---
                bank_summary_kes["change_KES"] = bank_summary_kes["closing_balance_KES"] - bank_summary_kes["opening_balance_KES"]
                bank_summary_kes["growth_%"] = (
                    (bank_summary_kes["change_KES"] / bank_summary_kes["opening_balance_KES"].replace(0, float("nan"))) * 100
                )

                # === 🧭 Visual 1: Total Cash Distribution by Bank ===
                fig_bank_dist = go.Figure(data=[
                    go.Pie(
                        labels=bank_summary_kes["bank_clean"],
                        values=bank_summary_kes["closing_balance_KES"],
                        hole=0.4,
                        textinfo="label+percent",
                        insidetextorientation="radial"
                    )
                ])
                fig_bank_dist.update_layout(
                    title_text="🏦 Total Cash Distribution by Bank (KES Equivalent)",
                    showlegend=True
                )
                st.plotly_chart(fig_bank_dist, use_container_width=True)

                st.markdown("""
                **Insight:**  
                - Each bank combines **all its currency accounts** into one total (in KES).  
                - Helps you see **total exposure per bank**, not per currency account.
                """)

                # === 🧭 Visual 2: Opening vs Closing per Bank ===
                fig_bank_balances = go.Figure()
                fig_bank_balances.add_trace(go.Bar(
                    x=bank_summary_kes["bank_clean"],
                    y=bank_summary_kes["opening_balance_KES"],
                    name="Opening (KES)",
                    marker_color="royalblue"
                ))
                fig_bank_balances.add_trace(go.Bar(
                    x=bank_summary_kes["bank_clean"],
                    y=bank_summary_kes["closing_balance_KES"],
                    name="Closing (KES)",
                    marker_color="seagreen"
                ))
                fig_bank_balances.update_layout(
                    barmode="group",
                    title="🏦 Opening vs. Closing Balances by Bank (KES Equivalent)",
                    xaxis_title="Bank",
                    yaxis_title="KES Amount",
                    legend_title="Balance Type"
                )
                st.plotly_chart(fig_bank_balances, use_container_width=True)

                # === 🧭 Visual 3: Growth or Decline % by Bank ===
                fig_bank_growth = go.Figure()
                fig_bank_growth.add_trace(go.Bar(
                    x=bank_summary_kes["bank_clean"],
                    y=bank_summary_kes["growth_%"],
                    text=bank_summary_kes["growth_%"].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else ""),
                    textposition="outside",
                    marker_color=bank_summary_kes["growth_%"].apply(lambda x: "seagreen" if x >= 0 else "crimson")
                ))
                fig_bank_growth.update_layout(
                    title="📊 Growth/Decline by Bank (%)",
                    xaxis_title="Bank",
                    yaxis_title="Growth %",
                    yaxis_tickformat=".1f",
                    showlegend=False
                )
                st.plotly_chart(fig_bank_growth, use_container_width=True)

                # === 📋 Detailed Table ===
                st.markdown("### 📋 Detailed Bank Summary (KES Equivalent)")
                st.dataframe(
                    bank_summary_kes.style.format({
                        "opening_balance_KES": "KSh{:,.2f}".format,
                        "closing_balance_KES": "KSh{:,.2f}".format,
                        "change_KES": "KSh{:,.2f}".format,
                        "growth_%": "{:.2f}%".format,
                    })
                    .applymap(lambda v: "color: green" if isinstance(v, (int, float)) and v > 0 else ("color: red" if isinstance(v, (int, float)) and v < 0 else ""))
                )

                st.markdown("""
                **Interpretation:**  
                - `"bank_clean"` merges all variants (e.g. `"NCBA USD"`, `"NCBA EUR"`) into `"NCBA"`.  
                - Balances are **converted to KES** and summed.  
                - Growth metrics show performance across all currency accounts combined.
                """)


elif page_selection == "Adjacements Reconciliation":
    st.title("Local & Foreign Adjacements Reconciliation App")
    if not st.session_state.bank_dfs: st.warning("Please go to 'Bank Statement Management' to upload and process bank statements first.")
    else: (st.session_state.df_matched_adjustments_local, st.session_state.df_matched_adjustments_foreign, st.session_state.df_unmatched_adjustments_local, st.session_state.df_unmatched_adjustments_foreign, st.session_state.df_unmatched_bank_records) = fx_reconciliation_app(st.session_state.bank_dfs)
elif page_selection == "FX Trade Reconciliation":
    st.title("FX Trade Reconciliation App")
    if not st.session_state.bank_dfs: st.warning("Please go to 'Bank Statement Management' to upload and process bank statements first.")
    else: (st.session_state.df_matched_counterparty, st.session_state.df_matched_choice, st.session_state.df_unmatched_counterparty, st.session_state.df_unmatched_choice, st.session_state.df_unmatched_bank_trade) = graphed_analysis_app(st.session_state.bank_dfs)

elif page_selection == "Intermediary Reconciliation":
    st.title("Intermediary Bank Reconciliation App")
    if not st.session_state.bank_dfs: 
        st.warning("Please go to 'Bank Statement Management' to upload and process bank statements first.")
    else: 
        (st.session_state.df_matched_intermediary_credit,
         st.session_state.df_matched_intermediary_debit, 
         st.session_state.df_unmatched_intermediary_credit, 
         st.session_state.df_unmatched_intermediary_debit, 
         st.session_state.df_unmatched_bank_intermediary) = intermediary_bank_reconciliation_app(st.session_state.bank_dfs)

elif page_selection == "Interfund Reconciliation":
    st.title("Interfund Bank Reconciliation App")
    if not st.session_state.bank_dfs: 
        st.warning("Please go to 'Bank Statement Management' to upload and process bank statements first.")
    else: 
        (st.session_state.df_matched_interfund,
         st.session_state.df_unmatched_interfund) = interfund_bank_reconciliation_app(st.session_state.bank_dfs)

elif page_selection == "Business FX Reconciliation":
    st.title("Business FX Reconciliation App")
    if not st.session_state.bank_dfs: st.warning("Please go to 'Bank Statement Management' to upload and process bank statements first.")
    else:  business_reconciliation_app(st.session_state.df_matched_counterparty, st.session_state.df_matched_choice, debug_mode=st.session_state.debug_mode )
    
elif page_selection == "Cross-Match Analysis":
    st.title("Cross-Match Analysis")
    st.write("This section combines and compares the results from all reconciliation applications to find potential missed matches.")
    if (st.session_state.df_matched_adjustments_local.empty and 
        st.session_state.df_matched_adjustments_foreign.empty and 
        st.session_state.df_matched_counterparty.empty and 
        st.session_state.df_matched_choice.empty and
        st.session_state.df_matched_intermediary_credit.empty and
        st.session_state.df_matched_intermediary_debit.empty and
        st.session_state.df_matched_interfund.empty):  # NEW: Add interfund check
        st.warning("Please first run the 'Adjustments Reconciliation', 'FX Trade Reconciliation', 'Intermediary Bank Reconciliation', and 'Interfund Bank Reconciliation' apps to populate the dataframes needed for cross-matching.")
    else:
        if st.button("Perform Cross-Match Analysis"):
            with st.spinner("Performing cross-match analysis..."):
                run_cross_match_analysis(
                    st.session_state.df_matched_adjustments_local,
                    st.session_state.df_matched_adjustments_foreign,
                    st.session_state.df_matched_counterparty,
                    st.session_state.df_matched_choice,
                    st.session_state.df_matched_intermediary_credit,
                    st.session_state.df_matched_intermediary_debit,
                    st.session_state.df_matched_interfund,  # NEW: Add interfund matches
                    st.session_state.bank_dfs,
                    debug_mode=st.session_state.debug_mode
                )
        else: 
            st.info("Click the button above to run the cross-match analysis.")
        cross_match_analysis_app()