# main_dashboard.py (COMPLETE - Full Auto Column Mapping Restored with Change Tracking)
import streamlit as st
from io import BytesIO
import pandas as pd
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go
import plotly.express as px

# Import authentication system
from auth_system import (
    get_active_version_id, init_auth_db, require_auth, 
    require_role, user_management_ui, 
     date_based_retrieval_ui, 
    get_all_versions, log_audit
)

# Import UI components
from ui_components import (
    apply_theme, kpi_metric, empty_state, loading_state,
    ag_grid_table, fin_tabs, section_header, subsection_header,
    COLORS
)

# Import analytics dashboard
from analytics_dashboard import analytics_dashboard_modal, show_analytics_for_date

# Import functions from other pages
from fx_reconcilliation_app_page import fx_reconciliation_app
from intermediary_bank_reconciliation_page import intermediary_bank_reconciliation_app
from interfund_bank_reconciliation_page import interfund_bank_reconciliation_app
from fx_trade_reconciliation_page import graphed_analysis_app
from combine_match_results_page import run_cross_match_analysis, cross_match_analysis_app
from business_fx_reconciliation_page import business_reconciliation_app
from mpesa_reconciliation_app_page import mpesa_gl_reconciliation_app

import requests
import time
import re
from typing import Optional, Dict
import json
import hashlib

# Initialize authentication database
init_auth_db()


# ==================== CHANGE TRACKING SYSTEM ====================
class DataFrameChangeTracker:
    """Tracks changes to DataFrames to enable incremental saving"""
    
    def __init__(self):
        # Store hashes of DataFrames when they were last saved
        self.saved_hashes = {}
        # Store current hashes for comparison
        self.current_hashes = {}
        # Track which DataFrames have been modified
        self.modified_dfs = set()
    
    def _calculate_hash(self, df: pd.DataFrame) -> str:
        """Calculate a hash of a DataFrame to detect changes"""
        if df is None or df.empty:
            return "empty"
        # Use a combination of shape and content hash
        # Convert DataFrame to string representation and hash it
        df_str = df.to_csv(index=False).encode('utf-8')
        return hashlib.md5(df_str).hexdigest()
    
    def register_df(self, name: str, df: pd.DataFrame, mark_saved: bool = True):
        """Register a DataFrame for change tracking"""
        if df is not None and not df.empty:
            current_hash = self._calculate_hash(df)
            self.current_hashes[name] = current_hash
            if mark_saved:
                self.saved_hashes[name] = current_hash
                if name in self.modified_dfs:
                    self.modified_dfs.remove(name)
    
    def check_changes(self, name: str, df: pd.DataFrame) -> bool:
        """Check if a DataFrame has changed since last save"""
        if df is None or df.empty:
            return False
        
        current_hash = self._calculate_hash(df)
        self.current_hashes[name] = current_hash
        
        if name not in self.saved_hashes:
            # Never saved before - considered modified
            self.modified_dfs.add(name)
            return True
        
        if self.saved_hashes[name] != current_hash:
            self.modified_dfs.add(name)
            return True
        
        return False
    
    def mark_saved(self, name: str):
        """Mark a DataFrame as saved"""
        if name in self.current_hashes:
            self.saved_hashes[name] = self.current_hashes[name]
        if name in self.modified_dfs:
            self.modified_dfs.remove(name)
    
    def get_modified_dfs(self) -> set:
        """Get all DataFrames that have been modified"""
        return self.modified_dfs.copy()
    
    def get_unmodified_dfs(self) -> set:
        """Get all DataFrames that haven't been modified since last save"""
        tracked = set(self.saved_hashes.keys())
        return tracked - self.modified_dfs
    
    def get_all_tracked_dfs(self) -> set:
        """Get all tracked DataFrames"""
        return set(self.saved_hashes.keys())
    
    def has_changes(self) -> bool:
        """Check if any tracked DataFrame has changed"""
        return len(self.modified_dfs) > 0
    
    def get_change_summary(self) -> dict:
        """Get a summary of changes"""
        return {
            'total_tracked': len(self.saved_hashes),
            'modified': list(self.modified_dfs),
            'unmodified': list(self.get_unmodified_dfs())
        }

# --- Modern Sidebar CSS with Dark Mode Support ---
def get_sidebar_css(is_dark: bool):
    """Return sidebar CSS based on theme"""
    if is_dark:
        return """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
        * { font-family: 'Inter', sans-serif; }
        .main .block-container { padding-top: 1rem; padding-left: 1.5rem; padding-right: 1.5rem; }
        section[data-testid="stSidebar"] {
            background: linear-gradient(180deg, #1e293b 0%, #0f172a 100%);
            border-right: 1px solid #334155;
            box-shadow: 4px 0 20px rgba(0, 0, 0, 0.3);
        }
        section[data-testid="stSidebar"] .stMarkdown,
        section[data-testid="stSidebar"] .stSelectbox label,
        section[data-testid="stSidebar"] .stRadio label { color: #e2e8f0; }
        .sidebar-logo { text-align: center; padding: 1.5rem 1rem; margin-bottom: 1rem; border-bottom: 1px solid #334155; }
        .sidebar-logo-icon { font-size: 2.5rem; animation: float 3s ease-in-out infinite; display: inline-block; }
        .sidebar-logo-text { font-weight: 800; font-size: 1.25rem; background: linear-gradient(135deg, #a78bfa, #c4b5fd); -webkit-background-clip: text; -webkit-text-fill-color: transparent; margin-top: 0.5rem; }
        .sidebar-logo-sub { font-size: 0.7rem; color: #94a3b8; margin-top: 0.25rem; }
        .user-profile { background: linear-gradient(135deg, #334155, #1e293b); margin: 1rem; padding: 1rem; border-radius: 1rem; text-align: center; border: 1px solid #475569; }
        .user-avatar { width: 50px; height: 50px; background: linear-gradient(135deg, #a78bfa, #c4b5fd); border-radius: 50%; display: flex; align-items: center; justify-content: center; margin: 0 auto 0.75rem; font-size: 1.25rem; font-weight: 600; color: #1e293b; box-shadow: 0 4px 10px rgba(0, 0, 0, 0.3); }
        .user-name { font-weight: 600; color: #e2e8f0; font-size: 0.9rem; }
        .user-role { font-size: 0.7rem; color: #94a3b8; margin-top: 0.25rem; }
        .sidebar-section-header { font-size: 0.7rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; color: #64748b; padding: 0.5rem 1rem; margin-top: 0.5rem; }
        .sidebar-divider { height: 1px; background: linear-gradient(90deg, transparent, #334155, transparent); margin: 1rem; }
        .stRadio > div { gap: 0.25rem; }
        .stRadio label { display: flex; align-items: center; gap: 0.75rem; padding: 0.625rem 1rem; margin: 0.25rem 0; border-radius: 0.5rem; transition: all 0.3s ease; cursor: pointer; font-weight: 500; font-size: 0.875rem; color: #cbd5e1; }
        .stRadio label:hover { background: #334155; transform: translateX(4px); color: #a78bfa; }
        footer { display: none; }
        @keyframes float { 0%, 100% { transform: translateY(0px); } 50% { transform: translateY(-5px); } }
        @keyframes fadeIn { from { opacity: 0; transform: translateY(20px); } to { opacity: 1; transform: translateY(0); } }
        .fade-in { animation: fadeIn 0.5s ease-out; }
        </style>
        """
    else:
        return """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
        * { font-family: 'Inter', sans-serif; }
        .main .block-container { padding-top: 1rem; padding-left: 1.5rem; padding-right: 1.5rem; }
        section[data-testid="stSidebar"] {
            background: linear-gradient(180deg, #FFFFFF 0%, #F8F6FC 100%);
            border-right: none;
            box-shadow: 4px 0 20px rgba(75, 45, 143, 0.06);
        }
        section[data-testid="stSidebar"] .stMarkdown,
        section[data-testid="stSidebar"] .stSelectbox label,
        section[data-testid="stSidebar"] .stRadio label { color: #4B5563; }
        .sidebar-logo { text-align: center; padding: 1.5rem 1rem; margin-bottom: 1rem; border-bottom: 1px solid #EDE8F5; }
        .sidebar-logo-icon { font-size: 2.5rem; animation: float 3s ease-in-out infinite; display: inline-block; }
        .sidebar-logo-text { font-weight: 800; font-size: 1.25rem; background: linear-gradient(135deg, #4B2D8F, #6B4DB5); -webkit-background-clip: text; -webkit-text-fill-color: transparent; margin-top: 0.5rem; }
        .sidebar-logo-sub { font-size: 0.7rem; color: #6B7280; margin-top: 0.25rem; }
        .user-profile { background: linear-gradient(135deg, #F0EBF9, #E8E0F5); margin: 1rem; padding: 1rem; border-radius: 1rem; text-align: center; border: 1px solid #E2D9F3; }
        .user-avatar { width: 50px; height: 50px; background: linear-gradient(135deg, #4B2D8F, #6B4DB5); border-radius: 50%; display: flex; align-items: center; justify-content: center; margin: 0 auto 0.75rem; font-size: 1.25rem; font-weight: 600; color: white; box-shadow: 0 4px 10px rgba(75, 45, 143, 0.2); }
        .user-name { font-weight: 600; color: #4B2D8F; font-size: 0.9rem; }
        .user-role { font-size: 0.7rem; color: #6B7280; margin-top: 0.25rem; }
        .sidebar-section-header { font-size: 0.7rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; color: #9CA3AF; padding: 0.5rem 1rem; margin-top: 0.5rem; }
        .sidebar-divider { height: 1px; background: linear-gradient(90deg, transparent, #EDE8F5, transparent); margin: 1rem; }
        .stRadio > div { gap: 0.25rem; }
        .stRadio label { display: flex; align-items: center; gap: 0.75rem; padding: 0.625rem 1rem; margin: 0.25rem 0; border-radius: 0.5rem; transition: all 0.3s ease; cursor: pointer; font-weight: 500; font-size: 0.875rem; }
        .stRadio label:hover { background: #F0EBF9; transform: translateX(4px); }
        footer { display: none; }
        @keyframes float { 0%, 100% { transform: translateY(0px); } 50% { transform: translateY(-5px); } }
        @keyframes fadeIn { from { opacity: 0; transform: translateY(20px); } to { opacity: 1; transform: translateY(0); } }
        .fade-in { animation: fadeIn 0.5s ease-out; }
        </style>
        """

# --- FastForex API Configuration ---
FASTFOREX_API_URL = "https://api.fastforex.io/fetch-all"
FASTFOREX_API_KEY = "4b744777d6-9c3eed3143-t4gxsb"
TARGET_CURRENCIES = ["KES", "USD", "EUR", "GBP", "CNY", "UGX", "RWF", "TZS", "ZAR"]
CACHE_FILE = "exchange_rates_cache.json"

# --- Currency Code Mapping ---
CURRENCY_NAME_MAP = {
    "KES": "KENYA SHILLING",
    "USD": "US DOLLAR",
    "GBP": "STG POUND", 
    "EUR": "EURO",
    "CNY": "CHINESE YUAN",
    "UGX": "UGANDA SHILLING",
    "RWF": "RWANDA FRANC",
    "TZS": "TANZANIA SHILLING",
    "ZAR": "SA RAND"
}

def get_live_exchange_rates(base_currency: str = "KES") -> Optional[Dict[str, float]]:
    """Fetch live exchange rates from FastForex API."""
    try:
        params = {"from": base_currency, "api_key": FASTFOREX_API_KEY}
        headers = {"accept": "application/json", "User-Agent": "Mozilla/5.0"}
        response = requests.get(FASTFOREX_API_URL, params=params, headers=headers, timeout=30)
        
        if response.status_code != 200:
            return None

        data = response.json()
        if "results" not in data:
            return None

        results = data.get('results', {})
        exchange_rates_to_kes = {}
        
        for currency_code, rate_from_kes in results.items():
            if currency_code in TARGET_CURRENCIES and rate_from_kes > 0:
                rate_to_kes = 1.0 / rate_from_kes
                exchange_rates_to_kes[currency_code] = rate_to_kes
        
        exchange_rates_to_kes["KES"] = 1.0
        return exchange_rates_to_kes
        
    except Exception:
        return get_fallback_rates()

def get_fallback_rates() -> Dict[str, float]:
    """Provide fallback exchange rates."""
    return {"KES": 1.0, "USD": 129.24, "EUR": 150.64, "GBP": 173.38, "CNY": 18.15, 
            "UGX": 0.037, "RWF": 0.089, "TZS": 0.053, "ZAR": 7.45}

def convert_to_kes(amount: float, currency: str, exchange_rates: Dict[str, float]) -> float:
    """Convert amount from given currency to KES."""
    if not amount or pd.isna(amount):
        return 0.0
    currency = currency.upper().strip()
    if currency == "KES":
        return float(amount)
    if currency in exchange_rates:
        return float(amount) * exchange_rates[currency]
    return float(amount)

# --- Constants and Global Mappings ---
DATE_FORMATS = [
    '%Y-%m-%d', '%Y/%m/%d', '%d.%m.%Y', '%Y.%m.%d', '%d/%m/%Y',
    '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%d.%m.%Y %H:%M:%S',
    '%Y.%m.%d %H:%M:%S', '%d/%m/%Y %H:%M:%S'
]

PREDEFINED_BANK_CURRENCY_OPTIONS = [
    "Absa KES", "Absa USD", "Absa EUR", "Absa GBP", "ABSA KES-SPECIAL", "ABSA USD-SPECIAL", 
    "ABSA EUR-SPECIAL", "ABSA GBP-SPECIAL", "ABSA Bank USD-DCD",
    "CBK KES", "CBK USD", "CBK EUR", "CBK GBP", "CBK UGX", "CBK TZS", "CBK RWF", "CBK ZAR", "CBK CNY",
    "Equity KES", "Equity USD", "Equity EUR", "Equity GBP",
    "I&M KES", "I&M USD", "I&M EUR", "I&M GBP",
    "KCB KES", "KCB USD", "KCB EUR", "KCB GBP", "KCB GBP - 1343013054",
    "Kingdom KES", "Kingdom USD", "Kingdom EUR", "Kingdom GBP",
    "NCBA KES", "NCBA USD", "NCBA EUR", "NCBA GBP",
    "SBM KES", "SBM USD", "SBM EUR", "SBM GBP",
    "UBA KES", "UBA USD", "UBA EUR", "UBA GBP",
    "BAAS Temporary KES", "BAAS Temporary USD", "BAAS Temporary EUR", "BAAS Temporary GBP",
    "FX Temporary KES", "FX Temporary USD", "FX Temporary EUR", "FX Temporary GBP",
    "Other Temporary KES", "Other Temporary USD", "Other Temporary EUR", "Other Temporary GBP",
    "Unclaimed Funds KES", "Unclaimed Funds USD", "Unclaimed Funds EUR", "Unclaimed Funds GBP",
    "Yeepay KES", "Yeepay USD", "Yeepay EUR", "Yeepay GBP", "Yeepay CNY"
]

FX_EXPECTED_COLUMNS = {
    'Amount': 'Amount', 'Operation': 'Operation', 'Completed At': 'Completed At',
    'Intermediary Account': 'Intermediary Account', 'Currency': 'Currency', 'Status': 'Status'
}

BANK_EXPECTED_COLUMNS = {
    'Date': ['Date', 'Transaction Date', 'Value Date', 'Value date'],
    'Credit': ['Credit', 'Credit Amount', 'Money In', 'Deposit', 'Credit amount'],
    'Debit': ['Debit', 'Debit Amount', 'Money Out', 'Withdrawal', 'Debit amount'],
    'Description': ['Description', 'Narrative', 'Transaction Details', 'Customer reference', 'Transaction Remarks:', 'Transaction Details', 'TransactionDetails', 'Transaction\nDetails'],
    'Running Balances': ['Running Balances', 'Running Balance', 'Running Balance (KES)', 'Running Balance (USD)', 'Running Balance (EUR)', 'Running Balance (GBP)', 'RUNNING BALANCES', 'RUNNING BALANCE', 'RUNNING BALANCE (KES)', 'RUNNING BALANCE (USD)', 'RUNNING BALANCE (EUR)', 'RUNNING BALANCE (GBP)']
}

# --- Helper Functions ---
def parse_date(date_str_raw):
    if pd.isna(date_str_raw) or date_str_raw == pd.NaT: return None
    if isinstance(date_str_raw, datetime): return date_str_raw
    if not isinstance(date_str_raw, str): date_str_raw = str(date_str_raw)
    date_str = date_str_raw.partition(" ")[0].strip() if " " in date_str_raw.strip() else date_str_raw.strip()
    for fmt in DATE_FORMATS:
        try: return datetime.strptime(date_str, fmt)
        except ValueError: continue
    return None

def safe_float(x):
    if pd.isna(x) or x is None: return None
    try:
        cleaned_x = str(x).replace(',', '').strip()
        return float(cleaned_x)
    except (ValueError, TypeError): return None

def process_uploaded_file(uploaded_file, sheet_name=None):
    uploaded_file.seek(0)
    if uploaded_file.name.endswith('.csv'):
        encodings = ['utf-8', 'utf-8-sig', 'latin1', 'ISO-8859-1', 'windows-1252']
        for enc in encodings:
            try:
                df = pd.read_csv(uploaded_file, encoding=enc)
                return df
            except Exception: continue
        st.error(f"Failed to decode CSV file '{uploaded_file.name}'")
        return pd.DataFrame()
    elif uploaded_file.name.endswith(('.xlsx', '.xls')):
        try:
            if isinstance(sheet_name, list):
                dfs = pd.read_excel(uploaded_file, sheet_name=sheet_name)
                return dfs
            else:
                df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
                return df
        except Exception as e:
            st.error(f"Error reading Excel file: {e}")
            return pd.DataFrame() if not isinstance(sheet_name, list) else {}
    else:
        st.error("Unsupported file type")
        return pd.DataFrame()

def get_excel_sheet_names(uploaded_file):
    uploaded_file.seek(0)
    try:
        excel_file = pd.ExcelFile(uploaded_file)
        return excel_file.sheet_names
    except Exception as e:
        st.error(f"Error getting sheet names: {e}")
        return []
    
def detect_column_type(series):
    if pd.api.types.is_datetime64_any_dtype(series):
        return 'datetime'
    if pd.api.types.is_numeric_dtype(series):
        return 'numeric'
    sample_size = min(100, len(series))
    sample = series.head(sample_size).dropna()
    if len(sample) == 0:
        return 'unknown'
    try:
        test_parsed = pd.to_datetime(sample, errors='coerce')
        success_rate = (test_parsed.notna().sum() / len(sample)) * 100
        if success_rate > 80:
            return 'date_string'
        else:
            return 'general_string'
    except:
        return 'general_string'

# --- Generate Cash Summary Report ---
def generate_cash_summary_report(per_bank_df: pd.DataFrame):
    """Generate the cash summary report with currency conversion to KES."""
    
    if not per_bank_df.empty:
        # --- Normalize column names ---
        per_bank_df.columns = (
            per_bank_df.columns.str.strip()
            .str.replace(" ", "_")
            .str.lower()
        )

        # --- Ensure required columns exist ---
        required_cols = ["currency", "bank", "opening_balance", "closing_balance"]
        missing = [c for c in required_cols if c not in per_bank_df.columns]
        if missing:
            st.warning(f"⚠️ Missing required columns: {missing}")
            return

        # --- Combine SPECIAL and DCD currency variants ---
        currency_merge_map = {
            "KES-SPECIAL": "KES",
            "USD-SPECIAL": "USD",
            "USD-DCD": "USD",
            "EUR-SPECIAL": "EUR",
            "GBP-SPECIAL": "GBP",
        }
        per_bank_df["currency"] = (
            per_bank_df["currency"].str.upper().replace(currency_merge_map)
        )

        # --- Get live exchange rates ---
        st.info("🔄 Fetching live exchange rates...")
        exchange_rates = get_live_exchange_rates("KES")
        
        # Display current rates for transparency
        if exchange_rates:
            rate_info_kes_to_foreign = " | ".join([f"1 KES = {1/rate:.4f} {curr}" 
                                for curr, rate in exchange_rates.items() 
                                if curr in ['USD', 'EUR', 'GBP'] and curr != "KES"])
            
            rate_info_foreign_to_kes = " | ".join([f"1 {curr} = {rate:.2f} KES" 
                                for curr, rate in exchange_rates.items() 
                                if curr in ['USD', 'EUR', 'GBP'] and curr != "KES"])
            
            st.caption(f"💱 Live Rates (KES to Foreign): {rate_info_kes_to_foreign}")
            st.caption(f"💱 Live Rates (Foreign to KES): {rate_info_foreign_to_kes}")

        
        # --- Compute currency summary automatically ---
        currency_summary = (
            per_bank_df.groupby("currency", as_index=False)[["opening_balance", "closing_balance"]]
            .sum()
            .sort_values("currency")
        )

        # --- Add KES conversion columns ---
        if exchange_rates:
            currency_summary['opening_balance_kes'] = currency_summary.apply(
                lambda x: convert_to_kes(x['opening_balance'], x['currency'], exchange_rates), 
                axis=1
            )
            currency_summary['closing_balance_kes'] = currency_summary.apply(
                lambda x: convert_to_kes(x['closing_balance'], x['currency'], exchange_rates), 
                axis=1
            )

        # --- Add Grand Total row ---
        grand_total_data = {
            "currency": "GRAND TOTAL",
            "opening_balance": currency_summary["opening_balance"].sum(),
            "closing_balance": currency_summary["closing_balance"].sum()
        }
        
        if 'opening_balance_kes' in currency_summary.columns:
            grand_total_data["opening_balance_kes"] = currency_summary["opening_balance_kes"].sum()
            grand_total_data["closing_balance_kes"] = currency_summary["closing_balance_kes"].sum()

        grand_total = pd.DataFrame([grand_total_data])
        currency_summary = pd.concat([currency_summary, grand_total], ignore_index=True)

        # === Create Bank Consolidated Summary (KES) ===
        per_bank_df["bank_clean"] = (
            per_bank_df["bank"]
            .astype(str)
            .str.replace(r"\b(USD|EUR|GBP|KES|CNY|ZAR|TZS|UGX|RWF)\b", "", regex=True)
            .str.replace(r"[-_/]+$", "", regex=True)
            .str.strip()
        )

        if exchange_rates:
            per_bank_df["fx_rate_to_KES"] = per_bank_df["currency"].map(exchange_rates).fillna(1.0)
            per_bank_df["opening_balance_KES"] = per_bank_df["opening_balance"] * per_bank_df["fx_rate_to_KES"]
            per_bank_df["closing_balance_KES"] = per_bank_df["closing_balance"] * per_bank_df["fx_rate_to_KES"]
        else:
            fx_rates = {"KES": 1.0, "USD": 130.0, "EUR": 140.0, "GBP": 160.0}
            per_bank_df["fx_rate_to_KES"] = per_bank_df["currency"].map(fx_rates).fillna(1.0)
            per_bank_df["opening_balance_KES"] = per_bank_df["opening_balance"] * per_bank_df["fx_rate_to_KES"]
            per_bank_df["closing_balance_KES"] = per_bank_df["closing_balance"] * per_bank_df["fx_rate_to_KES"]

        bank_summary_kes = (
            per_bank_df.groupby("bank_clean", as_index=False)[["opening_balance_KES", "closing_balance_KES"]]
            .sum()
            .sort_values("closing_balance_KES", ascending=False)
        )

        # --- Display KPIs ---
        total_opening_kes = currency_summary[currency_summary['currency'] != 'GRAND TOTAL']['opening_balance_kes'].sum() if 'opening_balance_kes' in currency_summary.columns else 0
        total_closing_kes = currency_summary[currency_summary['currency'] != 'GRAND TOTAL']['closing_balance_kes'].sum() if 'closing_balance_kes' in currency_summary.columns else 0
        growth_pct = ((total_closing_kes - total_opening_kes) / total_opening_kes * 100) if total_opening_kes > 0 else 0
        
        
        # --- Charts Row ---
        col1, col2 = st.columns(2)
        
        with col1:
            if 'closing_balance_kes' in currency_summary.columns:
                chart_df = currency_summary[currency_summary['currency'] != 'GRAND TOTAL'].copy()
                fig = go.Figure(data=[
                    go.Bar(
                        x=chart_df['currency'],
                        y=chart_df['closing_balance_kes'],
                        marker_color='#4B2D8F',
                        text=chart_df['closing_balance_kes'].apply(lambda x: f'{x:,.0f}'),
                        textposition='outside'
                    )
                ])
                fig.update_layout(
                    title="Closing Balance by Currency (KES)",
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    height=400,
                    xaxis_title="Currency",
                    yaxis_title="KES Amount"
                )
                st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            if 'closing_balance_kes' in currency_summary.columns:
                chart_df = currency_summary[currency_summary['currency'] != 'GRAND TOTAL'].copy()
                fig = go.Figure(data=[
                    go.Pie(
                        labels=chart_df['currency'],
                        values=chart_df['closing_balance_kes'],
                        hole=0.4,
                        marker=dict(colors=['#4B2D8F', '#6B4DB5', '#8B6DC5', '#F5C842', '#7D1128'])
                    )
                ])
                fig.update_layout(
                    title="Currency Distribution (KES)",
                    height=400,
                    showlegend=True
                )
                st.plotly_chart(fig, use_container_width=True)
        
        # --- Display currency summary table ---
        subsection_header("Currency Summary")
        display_summary = currency_summary.copy()
        if 'opening_balance_kes' in display_summary.columns:
            for col in ['opening_balance', 'closing_balance', 'opening_balance_kes', 'closing_balance_kes']:
                if col in display_summary.columns:
                    display_summary[col] = display_summary[col].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "0.00")
        
        st.dataframe(display_summary, use_container_width=True)
        
        # --- Bank Consolidated Summary (KES) ---
        subsection_header("Bank Consolidated Summary (KES)")
        display_bank_summary = bank_summary_kes.copy()
        display_bank_summary["change_KES"] = display_bank_summary["closing_balance_KES"] - display_bank_summary["opening_balance_KES"]
        display_bank_summary["growth_%"] = (
            (display_bank_summary["change_KES"] / display_bank_summary["opening_balance_KES"].replace(0, float("nan"))) * 100
        )
        
        # Format for display
        for col in ['opening_balance_KES', 'closing_balance_KES', 'change_KES']:
            display_bank_summary[col] = display_bank_summary[col].apply(lambda x: f"KSh{x:,.2f}" if pd.notna(x) else "KSh0.00")
        display_bank_summary["growth_%"] = display_bank_summary["growth_%"].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "0.00%")
        
        st.dataframe(display_bank_summary, use_container_width=True)

        # === Create Excel workbook ===
        excel_buffer = BytesIO()
        wb = Workbook(write_only=False)
        ws = wb.active
        ws.title = "Cash Summary"

        kes_currency_format = '#,##0.00" KSh"'
        number_format = '#,##0.00'
        
        formats = {
            'kes_currency': kes_currency_format,
            'number': number_format,
            'header_bold': Font(bold=True, size=14),
            'section_header': Font(bold=True),
            'column_header': Font(bold=True),
            'grand_total': Font(bold=True, color="FFFFFF"),
        }

        report_date = pd.Timestamp.today().strftime("%d %B %Y").upper()
        ws.merge_cells("A1:L1")
        ws["A1"] = f"CASH SUMMARY AS AT {report_date}"
        ws["A1"].font = formats['header_bold']
        ws["A1"].alignment = Alignment(horizontal="center")

        currency_order = ["KES", "USD", "EUR", "GBP", "CNY", "UGX", "RWF", "TZS", "ZAR"]
        start_col = 1
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for i, currency in enumerate(currency_order):
            df_cur = per_bank_df[per_bank_df["currency"].str.upper() == currency]
            if df_cur.empty:
                continue

            col_offset = (i * 3) + start_col

            ws.merge_cells(start_row=3, start_column=col_offset, end_row=3, end_column=col_offset + 2)
            ws.cell(row=3, column=col_offset).value = f"BANK {currency} ACCOUNTS"
            ws.cell(row=3, column=col_offset).font = formats['section_header']
            ws.cell(row=3, column=col_offset).alignment = Alignment(horizontal="center")

            headers = ["BANK NAME", "OPENING BALANCE", "CLOSING BALANCE"]
            for j, header in enumerate(headers):
                cell = ws.cell(row=4, column=col_offset + j, value=header)
                cell.font = formats['column_header']
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            for r_idx, row in enumerate(df_cur.itertuples(index=False), start=5):
                row_dict = row._asdict()
                bank_name = row_dict.get("bank", "")
                opening = row_dict.get("opening_balance", 0)
                closing = row_dict.get("closing_balance", 0)

                ws.cell(row=r_idx, column=col_offset, value=bank_name)
                ws.cell(row=r_idx, column=col_offset + 1, value=opening)
                ws.cell(row=r_idx, column=col_offset + 2, value=closing)

                for j in range(3):
                    cell = ws.cell(row=r_idx, column=col_offset + j)
                    cell.border = thin_border
                    if j > 0:
                        cell.number_format = formats['number']
                        cell.alignment = Alignment(horizontal="right")

        # === Add Bank Consolidated Summary (KES) on the same sheet ===
        consolidated_start_row = ws.max_row + 3
        
        ws.merge_cells(start_row=consolidated_start_row, start_column=1, end_row=consolidated_start_row, end_column=5)
        ws.cell(row=consolidated_start_row, column=1, value="BANK CONSOLIDATED SUMMARY (KES)").font = Font(bold=True, size=12)
        
        consolidated_headers = ["BANK", "OPENING BALANCE (KES)", "CLOSING BALANCE (KES)", "CHANGE (KES)", "GROWTH %"]
        for j, header in enumerate(consolidated_headers):
            cell = ws.cell(row=consolidated_start_row + 1, column=j + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for r_idx, row in enumerate(bank_summary_kes.itertuples(index=False), start=consolidated_start_row + 2):
            bank_name = row.bank_clean
            opening_kes = row.opening_balance_KES
            closing_kes = row.closing_balance_KES
            change_kes = closing_kes - opening_kes
            growth_pct = (change_kes / opening_kes * 100) if opening_kes != 0 else 0

            ws.cell(row=r_idx, column=1, value=bank_name)
            ws.cell(row=r_idx, column=2, value=opening_kes)
            ws.cell(row=r_idx, column=3, value=closing_kes)
            ws.cell(row=r_idx, column=4, value=change_kes)
            ws.cell(row=r_idx, column=5, value=growth_pct)

            for c in range(1, 6):
                cell = ws.cell(row=r_idx, column=c)
                cell.border = thin_border
                
                if c in [2, 3, 4]:
                    cell.number_format = kes_currency_format
                    cell.alignment = Alignment(horizontal="right")
                elif c == 5:
                    cell.number_format = "0.00%"
                    cell.alignment = Alignment(horizontal="right")

        # === Totals Section ===
        total_row_start = ws.max_row + 2
        
        if exchange_rates:
            rate_display = []
            for curr in ['USD', 'EUR', 'GBP', 'CNY', 'ZAR']:
                if curr in exchange_rates and curr != "KES":
                    rate_display.append(f"1 {curr} = {exchange_rates[curr]:.2f} KES")
            
            if rate_display:
                ws.merge_cells(start_row=total_row_start, start_column=1, end_row=total_row_start, end_column=6)
                rate_text = "Exchange Rates (to KES): " + " | ".join(rate_display)
                ws.cell(row=total_row_start, column=1, value=rate_text).font = Font(italic=True, size=9)
                total_row_start += 1

        ws.merge_cells(start_row=total_row_start, start_column=1, end_row=total_row_start, end_column=6)
        ws.cell(row=total_row_start, column=1, value="TOTALS BY CURRENCY").font = Font(bold=True, size=12)

        totals_headers = ["CURRENCY", "OPENING TOTAL", "CLOSING TOTAL"]
        if 'opening_balance_kes' in currency_summary.columns:
            totals_headers.extend(["OPENING (KES)", "CLOSING (KES)"])

        for j, header in enumerate(totals_headers):
            cell = ws.cell(row=total_row_start + 1, column=j + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for r_idx, row in enumerate(currency_summary.itertuples(index=False), start=total_row_start + 2):
            row_dict = row._asdict()
            currency = row_dict.get("currency", "")
            opening_total = row_dict.get("opening_balance", 0)
            closing_total = row_dict.get("closing_balance", 0)

            ws.cell(row=r_idx, column=1, value=currency)
            ws.cell(row=r_idx, column=2, value=opening_total)
            ws.cell(row=r_idx, column=3, value=closing_total)

            col_offset = 3
            if 'opening_balance_kes' in currency_summary.columns:
                opening_kes = row_dict.get("opening_balance_kes", 0)
                closing_kes = row_dict.get("closing_balance_kes", 0)
                
                ws.cell(row=r_idx, column=4, value=opening_kes)
                ws.cell(row=r_idx, column=5, value=closing_kes)
                col_offset = 5

            for c in range(1, col_offset + 1):
                cell = ws.cell(row=r_idx, column=c)
                cell.border = thin_border
                if c > 1:
                    if c >= 4:
                        cell.number_format = kes_currency_format
                    else:
                        cell.number_format = number_format
                    cell.alignment = Alignment(horizontal="right")

            if currency == "GRAND TOTAL":
                for c in range(1, col_offset + 1):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.font = formats['grand_total']
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for i, col_cells in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(i)
            for cell in col_cells:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    continue
            ws.column_dimensions[col_letter].width = max_length + 3

        wb.save(excel_buffer)
        excel_buffer.seek(0)

        st.download_button(
            label="⬇️ Download Cash Summary Excel Report",
            data=excel_buffer,
            file_name=f"Cash_Summary_{report_date.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        
        st.success("✅ Cash summary report generated successfully!")

# --- Modern Header Component ---
def modern_header(title: str, subtitle: str = "", icon: str = "💱"):
    is_dark = st.session_state.get('dark_mode', False)
    text_color = "#e2e8f0" if is_dark else "#4B5563"
    
    st.markdown(f"""
    <div class="fade-in" style="margin-bottom: 2rem;">
        <div style="display: flex; align-items: center; gap: 1rem; margin-bottom: 0.5rem;">
            <div style="font-size: 2.5rem; animation: float 3s ease-in-out infinite;">{icon}</div>
            <div>
                <h1 style="font-size: 2rem; font-weight: 700; background: linear-gradient(135deg, #4B2D8F, #6B4DB5); -webkit-background-clip: text; -webkit-text-fill-color: transparent; margin: 0;">
                    {title}
                </h1>
                {f'<p style="color: {text_color}; margin: 0.25rem 0 0 0;">{subtitle}</p>' if subtitle else ''}
            </div>
        </div>
        <div style="height: 3px; background: linear-gradient(90deg, #4B2D8F, #F5C842, #4B2D8F); border-radius: 3px; width: 100%;"></div>
    </div>
    """, unsafe_allow_html=True)

# --- Theme Toggle Button ---
def sidebar_theme_toggle():
    is_dark = st.session_state.get('dark_mode', False)
    button_label = "🌙 Dark Mode" if not is_dark else "☀️ Light Mode"
    if st.button(button_label, use_container_width=True):
        st.session_state.dark_mode = not is_dark
        st.rerun()


# --- Function to update change tracker after reconciliation ---
def update_change_tracker_for_reconciliation():
    """Update the change tracker with all reconciliation DataFrames"""
    change_tracker = st.session_state.change_tracker
    
    reconciliation_keys = [
        # Original reconciliation keys
        'df_matched_adjustments_local', 'df_matched_adjustments_foreign',
        'df_unmatched_adjustments_local', 'df_unmatched_adjustments_foreign',
        'df_unmatched_bank_records', 'df_matched_counterparty',
        'df_matched_choice', 'df_unmatched_counterparty', 'df_unmatched_choice',
        'df_unmatched_bank_trade', 'df_matched_intermediary_credit',
        'df_matched_intermediary_debit', 'df_unmatched_intermediary_credit',
        'df_unmatched_intermediary_debit', 'df_unmatched_bank_intermediary',
        'df_matched_interfund', 'df_unmatched_interfund', 'merged_bank_statement',
        
        # Additional missing states from keys_to_clear
        'matched_local', 'matched_foreign', 'unmatched_local', 'unmatched_foreign',
        'bank_records', 'moved_local_matched', 'moved_local_unmatched',
        'moved_foreign_matched', 'moved_foreign_unmatched', 'moved_bank_records',
        'deleted_local_matched', 'deleted_local_unmatched', 'deleted_foreign_matched',
        'deleted_foreign_unmatched', 'deleted_bank_records', 'audit_moves_log',
        'audit_deletes_log', 'moved_stats', 'deleted_stats'
    ]
    
    # for key in reconciliation_keys:
    #     if key in st.session_state:
    #         df = st.session_state[key]
    #         if df is not None and not df.empty:
    #             print(f"Registering changes for {key} with {len(df)} records")
    #             # change_tracker.check_changes(key, df)


# --- Main App ---
def main():
    st.set_page_config(page_title="ChoiceBank FX Reconciliation Dashboard", layout="wide", page_icon="💱", initial_sidebar_state="expanded")
    
    if 'dark_mode' not in st.session_state:
        st.session_state.dark_mode = False
    
    apply_theme()
    st.markdown(get_sidebar_css(st.session_state.dark_mode), unsafe_allow_html=True)
    
    if 'authenticated' not in st.session_state or not st.session_state['authenticated']:
        from auth_system import login_ui
        login_ui()
        return
    
    with st.sidebar:
        st.markdown("""
        <div class="sidebar-logo">
            <div class="sidebar-logo-icon">💱</div>
            <div class="sidebar-logo-text">ChoiceBank</div>
            <div class="sidebar-logo-sub">FX Reconciliation Platform</div>
        </div>
        """, unsafe_allow_html=True)
        
        user = st.session_state['user']
        user_name = user.get('full_name', user.get('username', 'User'))
        user_role = user.get('role', 'viewer')
        user_initial = user_name[0].upper() if user_name else "U"
        
        st.markdown(f"""
        <div class="user-profile">
            <div class="user-avatar">{user_initial}</div>
            <div class="user-name">{user_name}</div>
            <div class="user-role">{user_role.title()} Role</div>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)
        st.markdown('<div class="sidebar-section-header">📋 MAIN MENU</div>', unsafe_allow_html=True)
        
        nav_options = {
            "📁 Bank Management": "Bank Statement Management",
            "🔄 Adjustments": "Adjustments Reconciliation",
            "💱 FX Trade": "FX Trade Reconciliation",
            "🏦 Intermediary": "Intermediary Reconciliation",
            "🔄 Interfund": "Interfund Reconciliation",
            "💼 Business FX": "Business FX Reconciliation",
            "🔗 Cross-Match": "Cross-Match Analysis",
            "📱 M-Pesa": "M-Pesa & GL Reconciliation",
            "📊 Analytics": "Analytics",
        }
        
        # Main area navigation with vertical buttons
        st.markdown("### Module Navigation")
        st.markdown("Select a module to continue:")

        selected_page = None
        for page_name in nav_options.keys():
            col1 = st.columns([1])
            with col1[0]:
                if st.button(
                    page_name,
                    key=f"main_nav_{page_name}",
                    use_container_width=True,
                    type="primary" if st.session_state.get('selected_page') == page_name else "secondary"
                ):
                    selected_page = page_name
                    st.session_state.selected_page = page_name

        if selected_page is None:
            selected_page = st.session_state.get('selected_page', list(nav_options.keys())[0])

        page_selection = nav_options[selected_page]
        st.markdown("---")
        
        st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)
        
        if st.session_state['user']['role'] == 'admin':
            st.markdown('<div class="sidebar-section-header">⚙️ ADMIN</div>', unsafe_allow_html=True)
            col1, col2 = st.columns(2)
            with col1:
                if st.button("👥 Users", use_container_width=True):
                    st.session_state['admin_page'] = 'users'
            with col2:
                if st.button("📌 Versions", use_container_width=True):
                    st.session_state['admin_page'] = 'versions'
            st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)
        
        st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)
        sidebar_theme_toggle()
        
        if st.button("🚪 Logout", use_container_width=True):
            from auth_system import logout_user
            logout_user(st.session_state['session_token'])
            st.session_state.clear()
            st.rerun()
        
        versions_df = get_all_versions()
        if not versions_df.empty:
            active_version = versions_df[versions_df['is_active'] == 1]
            if not active_version.empty:
                st.caption(f"📌 {active_version.iloc[0]['version_name']}")
    
    # Admin page handling
    if 'admin_page' in st.session_state:
        if st.session_state['admin_page'] == 'users' and st.session_state['user']['role'] == 'admin':
            modern_header("User Management", "Manage system users and roles", "👥")
            user_management_ui()
            if st.button("← Back to Main"):
                del st.session_state['admin_page']
                st.rerun()
            return

        elif st.session_state['admin_page'] == 'historical':
            modern_header("Historical Data", "View and load historical reconciliation data", "📅")
            date_based_retrieval_ui()
            if st.button("← Back to Main"):
                del st.session_state['admin_page']
                st.rerun()
            return

    # Initialize session state variables
    if 'bank_dfs' not in st.session_state:
        st.session_state.bank_dfs = {}
    if 'cached_bank_files' not in st.session_state:
        st.session_state.cached_bank_files = {}
    if 'raw_bank_data_previews' not in st.session_state:
        st.session_state.raw_bank_data_previews = {}
    if 'merged_bank_statement' not in st.session_state:
        st.session_state.merged_bank_statement = pd.DataFrame()
    if 'debug_mode' not in st.session_state:
        st.session_state.debug_mode = False
    # Add change tracker to session state
    if 'change_tracker' not in st.session_state:
        st.session_state.change_tracker = DataFrameChangeTracker()
    
    # --- Main App Logic ---
    if page_selection == "Bank Statement Management":
        modern_header("Bank Statement Management", "Upload and configure your bank statements. These statements will then be available for all reconciliation modules.", "📁")
        
        uploaded_files = st.file_uploader("Upload Bank Statement(s) (CSV/Excel)", type=["csv", "xlsx"], accept_multiple_files=True, key="bank_uploader_main")
        
        if uploaded_files:
            for file in uploaded_files:
                if file.name not in st.session_state.cached_bank_files:
                    file_bytes = file.read()
                    st.session_state.cached_bank_files[file.name] = {"content": file_bytes, "type": file.type}

        files_to_delete = []

        if st.session_state.cached_bank_files:
            st.markdown("### 📄 Uploaded Bank Statements:")
            for file_name, file_data in st.session_state.cached_bank_files.items():
                file_key = file_name.lower().replace('.', '_')

                with st.expander(f"🗂️ {file_name}", expanded=True):
                    col1, col2 = st.columns([8, 2])
                    with col1: st.markdown(f"**File Name:** `{file_name}`")
                    with col2:
                        if st.button("❌ Remove", key=f"remove_{file_name}"):
                            files_to_delete.append(file_name)
                            continue

                    if file_key not in st.session_state.raw_bank_data_previews:
                        fake_file = BytesIO(file_data["content"])
                        fake_file.name = file_name

                        if file_name.endswith('.xlsx'):
                            sheet_names = get_excel_sheet_names(fake_file)
                            selected_sheets = [sheet_names[0]] if sheet_names else []
                            if selected_sheets:
                                dfs = process_uploaded_file(fake_file, sheet_name=selected_sheets)
                                df_dict = dfs if isinstance(dfs, dict) else {selected_sheets[0]: dfs}
                            else:
                                df_dict = {}
                        else:
                            sheet_names = []
                            selected_sheets = []
                            df_dict = {"CSV": process_uploaded_file(fake_file)}

                        standardized_names = {sheet_name: "" for sheet_name in df_dict.keys()}
                        st.session_state.raw_bank_data_previews[file_key] = {
                            'file_obj': fake_file, 
                            'df_raw_dict': df_dict, 
                            'sheet_names': sheet_names,
                            'selected_sheets': selected_sheets,
                            'column_mappings': {}, 
                            'standardized_names': standardized_names
                        }

                    data = st.session_state.raw_bank_data_previews[file_key]
                    df_bank_raw_dict = data['df_raw_dict']

                    if file_name.endswith('.xlsx') and data['sheet_names']:
                        current_sheets = st.multiselect(
                            f"Select Sheets for {file_name}:", 
                            data['sheet_names'],
                            default=data['selected_sheets'],
                            key=f"bank_sheet_selector_{file_key}"
                        )
                        
                        if set(current_sheets) != set(data['selected_sheets']):
                            data['selected_sheets'] = current_sheets
                            if current_sheets:
                                fake_file = BytesIO(file_data["content"])
                                fake_file.name = file_name
                                dfs = process_uploaded_file(fake_file, sheet_name=current_sheets)
                                df_bank_raw_dict = dfs if isinstance(dfs, dict) else {current_sheets[0]: dfs}
                                for sheet_name, df in df_bank_raw_dict.items():
                                    if df is not None:
                                        df.columns = df.columns.str.strip()
                                
                                for sheet_name in df_bank_raw_dict.keys():
                                    if sheet_name not in data['standardized_names']:
                                        data['standardized_names'][sheet_name] = ""
                                
                                st.info(f"Selected {len(current_sheets)} sheet(s) for {file_name}.")
                            else:
                                df_bank_raw_dict = {}
                            data['df_raw_dict'] = df_bank_raw_dict

                    if df_bank_raw_dict:
                        for sheet_name, df_bank_raw in df_bank_raw_dict.items():
                            if df_bank_raw is not None and not df_bank_raw.empty:
                                st.markdown(f"---")
                                st.subheader(f"Sheet: {sheet_name}")
                                
                                # Date conversion
                                if sheet_name in data['df_raw_dict']:
                                    df_processed = data['df_raw_dict'][sheet_name].copy()
                                    conversion_log = []
                                    
                                    for col in df_processed.columns:
                                        col_type = detect_column_type(df_processed[col])
                                        if col_type in ['datetime']:
                                            try:
                                                original_sample = df_processed[col].head(3).tolist()
                                                df_processed[col] = pd.to_datetime(df_processed[col], errors='coerce')
                                                successful_conversions = df_processed[col].notna().sum()
                                                total_rows = len(df_processed)
                                                
                                                if successful_conversions > 0:
                                                    df_processed[col] = df_processed[col].dt.strftime('%m/%d/%Y')
                                                    conversion_log.append(f"✅ **{col}**: Converted {successful_conversions}/{total_rows} dates")
                                                    with st.expander(f"Show conversion samples for '{col}'"):
                                                        st.write("**Before:**", original_sample)
                                                        st.write("**After:**", df_processed[col].head(3).tolist())
                                                else:
                                                    df_processed[col] = data['df_raw_dict'][sheet_name][col]
                                            except Exception as e:
                                                df_processed[col] = data['df_raw_dict'][sheet_name][col]
                                    
                                    data['df_raw_dict'][sheet_name] = df_processed
                                    
                                    if conversion_log:
                                        with st.expander("Date Conversion Summary"):
                                            for log_entry in conversion_log:
                                                st.write(log_entry)
                                
                                # Standardized name selector
                                selected_standardized_name = st.selectbox(
                                    f"Select Standardized Name for '{sheet_name}':", 
                                    options=[""] + PREDEFINED_BANK_CURRENCY_OPTIONS,
                                    index=PREDEFINED_BANK_CURRENCY_OPTIONS.index(data['standardized_names'].get(sheet_name, "")) + 1 
                                    if data['standardized_names'].get(sheet_name, "") in PREDEFINED_BANK_CURRENCY_OPTIONS else 0,
                                    key=f"standardized_name_selector_{file_key}_{sheet_name}"
                                )
                                data['standardized_names'][sheet_name] = selected_standardized_name

                                st.write(f"**Preview - {sheet_name}:**")
                                display_df = data['df_raw_dict'][sheet_name]
                                st.dataframe(display_df.head(100), use_container_width=True)

                                # Column Mapping (FULL AUTO MAPPING)
                                available_columns = display_df.columns.tolist()
                                available_columns.insert(0, "")
                                
                                if sheet_name not in data['column_mappings']:
                                    data['column_mappings'][sheet_name] = {}
                                current_mappings = data['column_mappings'][sheet_name]

                                st.write(f"**Column Mapping - {sheet_name}:**")
                                col_map_cols = st.columns(2)
                                for expected_col, default_val_list in BANK_EXPECTED_COLUMNS.items():
                                    initial_selection = current_mappings.get(expected_col)
                                    if not initial_selection:
                                        for default_val in default_val_list:
                                            if default_val.strip() in [col.strip() for col in display_df.columns]:
                                                initial_selection = default_val
                                                break
                                    
                                    with col_map_cols[0]: st.markdown(f"**{expected_col}**")
                                    with col_map_cols[1]:
                                        mapped_col = st.selectbox(
                                            f"Map '{expected_col}' to ({sheet_name}):", options=available_columns,
                                            index=available_columns.index(initial_selection) if initial_selection and initial_selection in available_columns else 0,
                                            key=f"bank_map_{file_key}_{sheet_name}_{expected_col}",
                                            label_visibility="collapsed"
                                        )
                                        data['column_mappings'][sheet_name][expected_col] = mapped_col if mapped_col else None
                            else:
                                st.warning(f"No data loaded for sheet '{sheet_name}' in {file_name}.")
                    else:
                        st.error(f"Could not load data from {file_name}.")
        
        for file_name in files_to_delete:
            del st.session_state.cached_bank_files[file_name]
            file_key = file_name.lower().replace('.', '_')
            if file_key in st.session_state.raw_bank_data_previews:
                del st.session_state.raw_bank_data_previews[file_key]
            st.success(f"Removed {file_name}")

        if st.button("🚀 Process All Bank Statements", type="primary"):
            with st.spinner("Processing bank statements..."):
                st.session_state.bank_dfs = {}
                all_success = True
                dfs_to_concat = []
                st.session_state.running_balances_col = None

                for file_key, data in st.session_state.raw_bank_data_previews.items():
                    st.info(f"Processing '{data['file_obj'].name}'...")

                    sheet_dfs = []
                    for sheet_name, df_raw in data['df_raw_dict'].items():
                        if df_raw is None or df_raw.empty:
                            st.warning(f"Skipping empty sheet '{sheet_name}' in '{data['file_obj'].name}'")
                            continue

                        sheet_standardized_name = data['standardized_names'].get(sheet_name, "")
                        if not sheet_standardized_name:
                            st.error(f"Missing standardized name for sheet '{sheet_name}' in '{data['file_obj'].name}'")
                            all_success = False
                            continue

                        if sheet_standardized_name in st.session_state.bank_dfs:
                            st.error(f"Duplicate standardized name '{sheet_standardized_name}' detected for sheet '{sheet_name}'. Please choose a unique name for each sheet.")
                            all_success = False
                            continue

                        df_to_process = df_raw.copy()
                        
                        sheet_mappings = data['column_mappings'].get(sheet_name, {})
                        
                        renamed_cols = {}
                        for expected_col, mapped_col in sheet_mappings.items():
                            if mapped_col and mapped_col in df_to_process.columns:
                                renamed_cols[mapped_col] = expected_col
                        
                        if renamed_cols:
                            df_to_process.rename(columns=renamed_cols, inplace=True)
                        df_to_process.columns = df_to_process.columns.str.strip()
                        
                        required_cols = ['Date', 'Credit', 'Debit', 'Running Balances']
                        missing_cols = [col for col in required_cols if col not in df_to_process.columns]
                        if missing_cols:
                            st.error(f"Validation failed for sheet '{sheet_name}' in '{data['file_obj'].name}'. Missing columns: {', '.join(missing_cols)}.")
                            all_success = False
                            continue

                        df_to_process['Date'] = df_to_process['Date'].apply(parse_date)
                        invalid_dates_mask = df_to_process['Date'].isna()
                        if invalid_dates_mask.any():
                            num_errors = invalid_dates_mask.sum()
                            st.warning(f"Warning in sheet '{sheet_name}' of '{data['file_obj'].name}': {num_errors} invalid dates found. These rows will be dropped.")
                            df_to_process = df_to_process[~invalid_dates_mask].copy()

                        df_to_process['Credit'] = df_to_process['Credit'].apply(safe_float)
                        df_to_process['Debit'] = df_to_process['Debit'].apply(safe_float)
                        df_to_process['Running Balances'] = df_to_process['Running Balances'].apply(safe_float)

                        df_to_process["Matched"] = False
                        df_to_process['Bank'] = sheet_standardized_name
                        df_to_process['Source_Sheet'] = sheet_name
                        df_to_process['Source_File'] = data['file_obj'].name
                        
                        sheet_dfs.append(df_to_process)
                        st.success(f"Processed: {data['file_obj'].name} - Sheet '{sheet_name}' as '{sheet_standardized_name}'")

                    if sheet_dfs:
                        for df in sheet_dfs:
                            bank_name = df['Bank'].iloc[0] if not df.empty else None
                            if bank_name:
                                st.session_state.bank_dfs[bank_name] = df
                        dfs_to_concat.extend(sheet_dfs)
                    else:
                        st.error(f"No valid sheets found in '{data['file_obj'].name}'")
                        all_success = False

                if all_success and dfs_to_concat:
                    st.session_state.merged_bank_statement = pd.concat(dfs_to_concat, ignore_index=True)
                    st.success("✅ All bank statements processed and merged!")
                    
                    # Update change tracker for bank_dfs and merged_bank_statement
                    change_tracker = st.session_state.change_tracker
                    if st.session_state.bank_dfs:
                        bank_dfs_str = str({k: v.shape if hasattr(v, 'shape') else len(v) for k, v in st.session_state.bank_dfs.items()})
                        current_bank_hash = hashlib.md5(bank_dfs_str.encode()).hexdigest()
                        change_tracker.current_hashes['bank_dfs'] = current_bank_hash
                        if 'bank_dfs' in change_tracker.saved_hashes and change_tracker.saved_hashes['bank_dfs'] != current_bank_hash:
                            change_tracker.modified_dfs.add('bank_dfs')
                        elif 'bank_dfs' not in change_tracker.saved_hashes:
                            change_tracker.modified_dfs.add('bank_dfs')
                    
                    if not st.session_state.merged_bank_statement.empty:
                        change_tracker.check_changes('merged_bank_statement', st.session_state.merged_bank_statement)
                    
                    if not st.session_state.merged_bank_statement.empty:
                        df_bal = st.session_state.merged_bank_statement.copy()
                        rb_col = 'Running Balances'
                        
                        df_bal.rename(columns={'Date': 'date', 'Debit': 'debit', 'Credit': 'credit', 'Bank': 'bank'}, inplace=True)
                        df_bal["currency"] = df_bal["bank"].apply(lambda x: str(x).split()[-1].upper())
                        df_bal = df_bal.sort_values(by=['bank', 'date'])
                        
                        per_bank_rows = []
                        for bank_name, df_bank in df_bal.groupby("bank"):
                            df_bank = df_bank.sort_values("date").reset_index(drop=True)
                            if len(df_bank) == 0:
                                continue
                            first_row = df_bank.iloc[0]
                            last_row = df_bank.iloc[-1]
                            currency = str(bank_name).split()[-1].upper()
                            running_balance_first = first_row[rb_col] if pd.notna(first_row[rb_col]) else 0
                            debit_first = first_row["debit"] if pd.notna(first_row["debit"]) else 0
                            credit_first = first_row["credit"] if pd.notna(first_row["credit"]) else 0

                            opening_balance = running_balance_first - credit_first + debit_first
                            closing_balance = last_row[rb_col] if pd.notna(last_row[rb_col]) else 0

                            per_bank_rows.append({"Bank": bank_name, "Currency": currency, 
                                                "Opening Balance": round(opening_balance, 2) if pd.notna(opening_balance) else 0, 
                                                "Closing Balance": round(closing_balance, 2) if pd.notna(closing_balance) else 0})

                        per_bank_df = pd.DataFrame(per_bank_rows).sort_values(by=["Currency", "Bank"]).reset_index(drop=True)
                        
                        # Per-Bank Table
                        subsection_header("Per-Bank Opening & Closing Balances")
                        st.dataframe(per_bank_df, use_container_width=True)
                        
                        csv_per_bank = per_bank_df.to_csv(index=False).encode("utf-8")
                        st.download_button("⬇️ Download Per-Bank Balances CSV", data=csv_per_bank, file_name="per_bank_balances.csv", mime="text/csv")
                        
                        # Currency Summary
                        currency_summary = per_bank_df.groupby("Currency").agg({"Opening Balance": "sum", "Closing Balance": "sum"}).round(2).reset_index()
                        subsection_header("Opening & Closing Balance Summary by Currency")
                        st.dataframe(currency_summary, use_container_width=True)
                        
                        csv_summary = currency_summary.to_csv(index=False).encode("utf-8")
                        st.download_button("⬇️ Download Currency Summary CSV", data=csv_summary, file_name="currency_balance_summary.csv", mime="text/csv")

                        # Monthly Transaction Volume Chart
                        st.markdown("---")
                        subsection_header("Monthly Transaction Volume")
                        df_chart = st.session_state.merged_bank_statement.copy()
                        df_chart['YearMonth'] = pd.to_datetime(df_chart['Date']).dt.to_period('M').astype(str)
                        df_chart['Credit'] = pd.to_numeric(df_chart['Credit'], errors='coerce').fillna(0)
                        df_chart['Debit'] = pd.to_numeric(df_chart['Debit'], errors='coerce').fillna(0)
                        monthly_volume = df_chart.groupby(['Bank', 'YearMonth']).agg(
                            Total_Credit=('Credit', 'sum'),
                            Total_Debit=('Debit', 'sum')
                        ).reset_index()
                        st.bar_chart(monthly_volume, x='YearMonth', y=['Total_Credit', 'Total_Debit'], color=['#008000', '#FF0000'])
                        
                        # Merged Bank Statement
                        st.markdown("---")
                        section_header("Merged Bank Statement")
                        st.dataframe(st.session_state.merged_bank_statement, use_container_width=True)
                        csv = st.session_state.merged_bank_statement.to_csv(index=False).encode("utf-8")
                        st.download_button("⬇️ Download Merged Bank Statement as CSV", data=csv, file_name="merged_bank_statement.csv", mime="text/csv")

                        # Generate Cash Summary Report
                        generate_cash_summary_report(per_bank_df)
    
    elif page_selection == "Adjustments Reconciliation":
        modern_header("Adjustments Reconciliation", "Local & Foreign Adjustments Reconciliation", "🔄")
        with st.spinner("Running FX Reconciliation..."):
            # Call the function and unpack the results
            (st.session_state.matched_local,
            st.session_state.matched_foreign,
            st.session_state.unmatched_local,
            st.session_state.unmatched_foreign,
            st.session_state.bank_records,
            st.session_state.moved_local_matched,
            st.session_state.moved_local_unmatched,
            st.session_state.moved_foreign_matched,
            st.session_state.moved_foreign_unmatched,
            st.session_state.moved_bank_records,
            st.session_state.deleted_local_matched,
            st.session_state.deleted_local_unmatched,
            st.session_state.deleted_foreign_matched,
            st.session_state.deleted_foreign_unmatched,
            st.session_state.deleted_bank_records,
            st.session_state.audit_moves_log,
            st.session_state.audit_deletes_log,
            st.session_state.moved_stats,
            st.session_state.deleted_stats,
            st.session_state.df_matched_adjustments_local,
            st.session_state.df_unmatched_adjustments_local,
            st.session_state.df_matched_adjustments_foreign,
            st.session_state.df_unmatched_adjustments_foreign,
            st.session_state.df_unmatched_bank_records) = fx_reconciliation_app(st.session_state.bank_dfs)
            
            # Update change tracker after reconciliation
            update_change_tracker_for_reconciliation()
    
    elif page_selection == "FX Trade Reconciliation":
        modern_header("FX Trade Reconciliation", "Match FX trades with bank statements", "💱")
        # if not st.session_state.bank_dfs:
        #     empty_state("No bank statements loaded", "📁", "Please go to Bank Statement Management first")
        # else:
        with st.spinner("Running FX Trade Reconciliation..."):
            # In main_dashboard.py, line 1495
            (st.session_state.df_matched_counterparty,
            st.session_state.df_matched_choice,
            st.session_state.df_unmatched_counterparty,
            st.session_state.df_unmatched_choice,
            st.session_state.df_unmatched_bank_trade,
            st.session_state.matched_local,  # moved_buy_matched
            st.session_state.unmatched_local,  # moved_buy_unmatched  
            st.session_state.matched_foreign,  # moved_sell_matched
            st.session_state.unmatched_foreign,  # moved_sell_unmatched
            st.session_state.bank_records,  # moved_bank_records_trade
            st.session_state.deleted_local_matched,  # deleted_buy_matched
            st.session_state.deleted_local_unmatched,  # deleted_buy_unmatched
            st.session_state.deleted_foreign_matched,  # deleted_sell_matched
            st.session_state.deleted_foreign_unmatched,  # deleted_sell_unmatched
            st.session_state.deleted_bank_records,  # deleted_bank_trade
            st.session_state.audit_moves_log,  # audit_moves_log_trade
            st.session_state.audit_deletes_log,  # audit_deletes_log_trade
            st.session_state.moved_stats,  # moved_stats_trade (as DataFrame)
            st.session_state.deleted_stats,  # deleted_stats_trade (as DataFrame)
            st.session_state.df_matched_adjustments_local,
            st.session_state.df_unmatched_adjustments_local,
            st.session_state.df_matched_adjustments_foreign,
            st.session_state.df_unmatched_adjustments_foreign,
            st.session_state.df_unmatched_bank_records) = graphed_analysis_app(st.session_state.bank_dfs)
            
            # Update change tracker after reconciliation
            update_change_tracker_for_reconciliation()


    elif page_selection == "Intermediary Reconciliation":
        modern_header("Intermediary Bank Reconciliation", "Match intermediary bank transactions", "🏦")
        # if not st.session_state.bank_dfs:
        #     empty_state("No bank statements loaded", "📁", "Please go to Bank Statement Management first")
        # else:
        with st.spinner("Running Intermediary Reconciliation..."):
            (st.session_state.df_matched_intermediary_credit,
            st.session_state.df_matched_intermediary_debit, 
            st.session_state.df_unmatched_intermediary_credit, 
            st.session_state.df_unmatched_intermediary_debit, 
            st.session_state.df_unmatched_bank_intermediary) = intermediary_bank_reconciliation_app(st.session_state.bank_dfs)
            
            # Update change tracker after reconciliation
            update_change_tracker_for_reconciliation()
    
    elif page_selection == "Interfund Reconciliation":
        modern_header("Interfund Bank Reconciliation", "Match interfund transfers", "🔄")
        # if not st.session_state.bank_dfs:
        #     empty_state("No bank statements loaded", "📁", "Please go to Bank Statement Management first")
        # else:
        with st.spinner("Running Interfund Reconciliation..."):
            (st.session_state.df_matched_interfund,
                st.session_state.df_unmatched_interfund) = interfund_bank_reconciliation_app(st.session_state.bank_dfs)
            
            # Update change tracker after reconciliation
            update_change_tracker_for_reconciliation()
    
    elif page_selection == "Business FX Reconciliation":
        modern_header("Business FX Reconciliation", "Match business FX transactions", "💼")
        # if not st.session_state.bank_dfs:
        #     empty_state("No bank statements loaded", "📁", "Please go to Bank Statement Management first")
        # else:
        business_reconciliation_app(st.session_state.df_matched_counterparty, st.session_state.df_matched_choice, debug_mode=st.session_state.debug_mode)
    
    elif page_selection == "Cross-Match Analysis":
        modern_header("Cross-Match Analysis", "Combine and compare results from all reconciliation modules", "🔗")
        if (st.session_state.get('df_matched_adjustments_local', pd.DataFrame()).empty):
            empty_state("No reconciliation data found", "📭", "Please first run the reconciliation apps")
        else:
            if st.button("🚀 Perform Cross-Match Analysis", type="primary", use_container_width=True):
                with st.spinner("Performing cross-match analysis..."):
                    run_cross_match_analysis(
                        st.session_state.get('df_matched_adjustments_local', pd.DataFrame()),
                        st.session_state.get('df_matched_adjustments_foreign', pd.DataFrame()),
                        st.session_state.get('df_matched_counterparty', pd.DataFrame()),
                        st.session_state.get('df_matched_choice', pd.DataFrame()),
                        st.session_state.get('df_matched_intermediary_credit', pd.DataFrame()),
                        st.session_state.get('df_matched_intermediary_debit', pd.DataFrame()),
                        st.session_state.get('df_matched_interfund', pd.DataFrame()),
                        st.session_state.bank_dfs,
                        debug_mode=st.session_state.debug_mode
                    )
            cross_match_analysis_app()
    
    elif page_selection == "M-Pesa & GL Reconciliation":
        modern_header("M-Pesa & GL Reconciliation", "Match mobile money transactions", "📱")
        mpesa_gl_reconciliation_app()
    
    elif page_selection == "Analytics":
        modern_header("Analytics Dashboard", "Visual insights and performance metrics", "📊")
        analytics_dashboard_modal()

if __name__ == "__main__":
    main()